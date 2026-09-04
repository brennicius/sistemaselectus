from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
import sqlite3, os, math, io, json
import pandas as pd
from datetime import datetime, date

app = Flask(__name__)
app.secret_key = 'selectus_cozinha_2024'
app.config['TEMPLATES_AUTO_RELOAD'] = True
DB = os.path.join(os.path.dirname(__file__), 'cozinha.db')

CATEGORIAS = [
    ('Cuscuz',         'bg-warning text-dark'),
    ('Saladas gourmet','bg-success'),
    ('Frutas',         'bg-danger'),
    ('Sanduíches',     'bg-primary'),
    ('Sobremesas',     'bg-purple'),
    ('Base saladas',   'bg-secondary'),
    ('Produção',       'bg-dark text-white'),
    ('Poke',           'bg-info text-dark'),
    ('Saladas +P',     'bg-teal'),
]

def _to_br(value, decimals):
    """Converte float para formato brasileiro: 1234.56 → 1.234,56"""
    if value is None:
        return '—'
    s = f'{float(value):,.{decimals}f}'
    return s.replace(',', 'X').replace('.', ',').replace('X', '.')

@app.template_filter('brl')
def brl_filter(value, decimals=2):
    """Formata como moeda brasileira: R$ 1.234,56"""
    return 'R$ ' + _to_br(value, decimals)

@app.template_filter('num_br')
def num_br_filter(value, decimals=2):
    """Formata número no padrão brasileiro"""
    return _to_br(value, decimals)

# Known conversions between purchase unit and usage unit
CONV = {
    ('kg','g'): 1000, ('kg','mg'): 1000000, ('kg','kg'): 1,
    ('g','g'): 1,     ('g','mg'): 1000,
    ('L','ml'): 1000, ('L','cl'): 100, ('L','L'): 1,
    ('ml','ml'): 1,
    ('un','un'): 1,
}

def conv_factor(from_u, to_u):
    if not from_u or not to_u:
        return None
    if from_u == to_u:
        return 1
    return CONV.get((from_u.strip(), to_u.strip()))

def get_db():
    c = sqlite3.connect(DB)
    c.row_factory = sqlite3.Row
    c.execute("PRAGMA foreign_keys = ON")
    return c

def init_db():
    db = get_db()
    db.executescript('''
        CREATE TABLE IF NOT EXISTS insumos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            unidade_compra TEXT NOT NULL,
            preco_compra REAL,
            unidade_uso TEXT NOT NULL,
            fator_conversao REAL NOT NULL DEFAULT 1,
            qtd_por_embalagem REAL,
            unid_embalagem TEXT
        );
        CREATE TABLE IF NOT EXISTS produtos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            descricao TEXT,
            rendimento REAL DEFAULT 1,
            unidade_rendimento TEXT DEFAULT 'un'
        );
        CREATE TABLE IF NOT EXISTS ficha_tecnica (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            produto_id INTEGER NOT NULL,
            insumo_id INTEGER NOT NULL,
            quantidade REAL NOT NULL,
            FOREIGN KEY (produto_id) REFERENCES produtos(id) ON DELETE CASCADE,
            FOREIGN KEY (insumo_id) REFERENCES insumos(id)
        );
    ''')

    # Migrations
    for col in ['qtd_por_embalagem REAL', 'unid_embalagem TEXT',
                'produto_vinculado_id INTEGER']:
        try:
            db.execute(f'ALTER TABLE insumos ADD COLUMN {col}')
        except Exception:
            pass
    for col in ['categoria TEXT', 'valor_referencia REAL',
                "status_revisao TEXT DEFAULT 'nao_revisada'",
                "status_revisao_custo TEXT DEFAULT 'nao_revisado'",
                'data_revisao TEXT',
                'data_revisao_custo TEXT']:
        try:
            db.execute(f'ALTER TABLE produtos ADD COLUMN {col}')
        except Exception:
            pass
    for col in ["status_revisao_custo TEXT DEFAULT 'nao_revisado'",
                'data_revisao_custo TEXT',
                "status_revisao_nome TEXT DEFAULT 'nao_revisado'",
                'observacao TEXT']:
        try:
            db.execute(f'ALTER TABLE insumos ADD COLUMN {col}')
        except Exception:
            pass
    for col in ['observacao TEXT']:
        try:
            db.execute(f'ALTER TABLE produtos ADD COLUMN {col}')
        except Exception:
            pass
    db.executescript('''
        CREATE TABLE IF NOT EXISTS historico_precos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            insumo_id INTEGER NOT NULL,
            preco_anterior REAL,
            preco_novo REAL,
            data_registro TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS historico_custo_fichas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            produto_id INTEGER NOT NULL,
            custo_anterior REAL,
            custo_novo REAL NOT NULL,
            data_registro TEXT DEFAULT CURRENT_TIMESTAMP
        );
    ''')
    # Seed inicial: registra custo atual de cada ficha se tabela vazia
    if db.execute('SELECT COUNT(*) FROM historico_custo_fichas').fetchone()[0] == 0:
        from datetime import datetime as _dt
        _now = _dt.now().strftime('%Y-%m-%d %H:%M:%S')
        for p_row in db.execute('SELECT id FROM produtos').fetchall():
            pid = p_row[0]
            rows = db.execute('''SELECT ft.quantidade, i.preco_compra, i.fator_conversao, i.produto_vinculado_id
                FROM ficha_tecnica ft JOIN insumos i ON i.id=ft.insumo_id WHERE ft.produto_id=?''', (pid,)).fetchall()
            total = 0.0; ok = True
            for r in rows:
                if r[3]: ok = False; break
                if r[1] is not None: total += r[0] * r[1] / r[2]
                else: ok = False; break
            if ok and total > 0:
                db.execute('INSERT INTO historico_custo_fichas (produto_id, custo_anterior, custo_novo, data_registro) VALUES (?,?,?,?)',
                           (pid, None, round(total, 6), _now))
    db.commit()

    db.executescript('''
        CREATE TABLE IF NOT EXISTS requisicoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            data TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'aberto',
            observacao TEXT,
            criado_em TEXT DEFAULT CURRENT_TIMESTAMP,
            fechado_em TEXT
        );
        CREATE TABLE IF NOT EXISTS requisicao_itens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            requisicao_id INTEGER NOT NULL,
            insumo_id INTEGER NOT NULL,
            qtd_enviada REAL NOT NULL,
            qtd_retornada REAL,
            FOREIGN KEY (requisicao_id) REFERENCES requisicoes(id) ON DELETE CASCADE,
            FOREIGN KEY (insumo_id) REFERENCES insumos(id)
        );
    ''')
    db.commit()

    # Criar produto inicial apenas se ainda não existir nenhum produto
    if not db.execute('SELECT 1 FROM produtos').fetchone():

        db.execute("INSERT INTO produtos (nome,descricao,rendimento,unidade_rendimento) VALUES(?,?,?,?)",
                   ('Salada Teriaky +P', '240 g', 1, 'un'))
        pid = db.execute('SELECT last_insert_rowid()').fetchone()[0]

        for nome, qty in [
            ('Frango Marinado', 80), ('Base Salada +P', 1), ('Mussarela Búfala', 20),
            ('Parmesão', 10), ('Tomate Seco', 10), ('Gergelim', 3), ('Molho Tarê', 30),
            ('Embalagem Bowl para Salada', 1), ('Kit Talher', 1),
            ('Etiqueta de Validade', 1), ('Etiqueta 90x40 Sanduíches/Saladas/Cuscuz', 1),
        ]:
            row = db.execute('SELECT id FROM insumos WHERE nome=?', (nome,)).fetchone()
            if row:
                db.execute('INSERT INTO ficha_tecnica (produto_id,insumo_id,quantidade) VALUES(?,?,?)', (pid, row['id'], qty))

    # Registro de Produção
    db.executescript('''
        CREATE TABLE IF NOT EXISTS registros_producao (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            data TEXT NOT NULL,
            rodada TEXT NOT NULL,
            status TEXT DEFAULT 'planejado',
            observacoes TEXT,
            criado_em TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS registro_itens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            registro_id INTEGER NOT NULL REFERENCES registros_producao(id) ON DELETE CASCADE,
            ordem INTEGER DEFAULT 0,
            categoria TEXT,
            produto TEXT NOT NULL,
            qtd_planejada INTEGER,
            qtd_portugues INTEGER,
            qtd_amaro INTEGER,
            qtd_izabel INTEGER,
            qtd_agnus INTEGER,
            qtd_produzida INTEGER,
            concluido INTEGER DEFAULT 0,
            observacao TEXT
        );
    ''')
    db.commit()

    # Migrations para estoque central e rastreamento de consumo
    try:
        db.execute('ALTER TABLE insumos ADD COLUMN estoque_central REAL DEFAULT 0')
    except Exception:
        pass
    db.executescript('''
        CREATE TABLE IF NOT EXISTS requisicao_consumos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            requisicao_id INTEGER NOT NULL,
            item_id INTEGER NOT NULL,
            insumo_id INTEGER NOT NULL,
            data TEXT NOT NULL,
            qtd_consumida REAL NOT NULL DEFAULT 0,
            qtd_sobra REAL,
            obs TEXT,
            criado_em TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS insumo_ajuste_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            insumo_id INTEGER NOT NULL,
            data TEXT NOT NULL,
            estoque_cozinha_antes REAL,
            estoque_central_antes REAL,
            estoque_cozinha_depois REAL,
            estoque_central_depois REAL,
            obs TEXT
        );
    ''')
    db.commit()
    db.close()

def _recalcular_historico_fichas(db, insumo_id):
    """Grava histórico de custo de todas as fichas afetadas por mudança de preço de um insumo."""
    from datetime import datetime as _dt
    now = _dt.now().strftime('%Y-%m-%d %H:%M:%S')
    produtos = db.execute('SELECT DISTINCT produto_id FROM ficha_tecnica WHERE insumo_id=?', (insumo_id,)).fetchall()
    for row in produtos:
        pid = row[0]
        custo_novo, ok = custo_produto(pid)
        if not ok:
            continue
        custo_novo = round(custo_novo, 6)
        ultimo = db.execute('SELECT custo_novo FROM historico_custo_fichas WHERE produto_id=? ORDER BY id DESC LIMIT 1', (pid,)).fetchone()
        custo_ant = ultimo[0] if ultimo else None
        if custo_ant != custo_novo:
            db.execute('INSERT INTO historico_custo_fichas (produto_id, custo_anterior, custo_novo, data_registro) VALUES (?,?,?,?)',
                       (pid, custo_ant, custo_novo, now))

def custo_produto(produto_id, _visited=None):
    """Calcula custo total de um produto. Suporta insumos vinculados a fichas técnicas."""
    if _visited is None:
        _visited = set()
    if produto_id in _visited:
        return 0.0, False  # ciclo detectado
    _visited = _visited | {produto_id}

    db = get_db()
    rows = db.execute('''
        SELECT ft.quantidade,
               i.preco_compra, i.fator_conversao,
               i.produto_vinculado_id,
               p2.rendimento AS rend_vinc,
               COALESCE(ft.aproveitamento, 100) AS aproveitamento
        FROM ficha_tecnica ft
        JOIN insumos i ON i.id = ft.insumo_id
        LEFT JOIN produtos p2 ON p2.id = i.produto_vinculado_id
        WHERE ft.produto_id = ?''', (produto_id,)).fetchall()
    db.close()

    total, incompleto = 0.0, False
    for r in rows:
        aprov = (r['aproveitamento'] or 100) / 100
        if r['produto_vinculado_id']:
            sub, sub_ok = custo_produto(r['produto_vinculado_id'], _visited)
            if sub_ok:
                rend = r['rend_vinc'] or 1
                total += r['quantidade'] / aprov * (sub / rend) / r['fator_conversao']
            else:
                incompleto = True
        elif r['preco_compra'] is not None:
            total += r['quantidade'] / aprov * (r['preco_compra'] / r['fator_conversao'])
        else:
            incompleto = True
    return total, not incompleto

# ── Dashboard ───────────────────────────────────────────────
@app.route('/')
def index():
    db = get_db()
    n_ins  = db.execute('SELECT COUNT(*) FROM insumos').fetchone()[0]
    n_prod = db.execute('SELECT COUNT(*) FROM produtos').fetchone()[0]
    s_prec = db.execute('SELECT COUNT(*) FROM insumos WHERE preco_compra IS NULL AND produto_vinculado_id IS NULL').fetchone()[0]
    prods  = db.execute('SELECT * FROM produtos ORDER BY nome').fetchall()
    cats_count = db.execute('''SELECT COALESCE(categoria,'Sem categoria') as cat, COUNT(*) as n
                               FROM produtos GROUP BY cat ORDER BY cat''').fetchall()
    db.close()
    cards = [{'id': p['id'], 'nome': p['nome'], 'descricao': p['descricao'],
              **dict(zip(['custo','completo'], custo_produto(p['id'])))} for p in prods]
    cat_map = {nome: cls for nome, cls in CATEGORIAS}
    return render_template('index.html', n_ins=n_ins, n_prod=n_prod, s_prec=s_prec, cards=cards, cats_count=cats_count, cat_map=cat_map)

# ── Insumos ──────────────────────────────────────────────────
@app.route('/insumos/sem-uso')
def insumos_sem_uso():
    db = get_db()
    rows = db.execute('''
        SELECT i.* FROM insumos i
        WHERE i.id NOT IN (SELECT DISTINCT insumo_id FROM ficha_tecnica)
        ORDER BY i.nome
    ''').fetchall()
    db.close()
    return render_template('insumos_sem_uso.html', insumos=rows)

@app.route('/insumos')
def insumos_lista():
    db = get_db()
    rows = db.execute('SELECT * FROM insumos ORDER BY nome').fetchall()
    hist_rows = db.execute('''
        SELECT h.insumo_id, h.preco_anterior, h.preco_novo, h.data_registro
        FROM historico_precos h
        WHERE h.id = (SELECT MAX(id) FROM historico_precos h2 WHERE h2.insumo_id = h.insumo_id)
    ''').fetchall()
    mudancas = {r['insumo_id']: dict(r) for r in hist_rows}
    db.close()
    sem_preco_ids = {r['id'] for r in rows if not r['preco_compra'] and not r['produto_vinculado_id']}
    return render_template('insumos_lista.html', insumos=rows, sem_preco_ids=sem_preco_ids, mudancas=mudancas)

def _parse_insumo_form(d):
    nome  = d['nome'].strip()
    uc    = d['unidade_compra'].strip()
    uu    = d['unidade_uso'].strip()
    preco = float(d['preco_compra']) if d.get('preco_compra','').strip() else None
    qtd_emb  = float(d['qtd_por_embalagem']) if d.get('qtd_por_embalagem','').strip() else None
    unid_emb = d.get('unid_embalagem','').strip() or None
    vinc = int(d['produto_vinculado_id']) if d.get('produto_vinculado_id','').strip() else None
    fornecedor = d.get('fornecedor','').strip() or None
    aprov_str = d.get('aproveitamento','').strip()
    aproveitamento = float(aprov_str) if aprov_str else 100.0
    aproveitamento = max(1.0, min(100.0, aproveitamento))

    if qtd_emb and unid_emb:
        cf = conv_factor(unid_emb, uu)
        fator = qtd_emb * cf if cf else float(d.get('fator_conversao') or 1)
    else:
        fator = float(d.get('fator_conversao') or 1)

    return nome, uc, preco, uu, fator, qtd_emb, unid_emb, vinc, fornecedor, aproveitamento

@app.route('/insumos/conversao', methods=['POST'])
def insumo_conv_preview():
    """AJAX: return calculated fator and preco_uso for live preview."""
    d = request.json
    qtd  = float(d.get('qtd') or 0)
    ue   = d.get('ue', '')
    uu   = d.get('uu', '')
    preco = float(d.get('preco') or 0)
    cf = conv_factor(ue, uu)
    if cf and qtd:
        fator = qtd * cf
        pu = preco / fator if fator else None
        return jsonify(fator=fator, preco_uso=pu, cf=cf, ok=True)
    return jsonify(ok=False, cf=cf)

def _get_produtos_para_vincular():
    db = get_db()
    prods = db.execute('SELECT id, nome FROM produtos ORDER BY nome').fetchall()
    db.close()
    return prods

@app.route('/insumos/novo', methods=['GET','POST'])
def insumo_novo():
    if request.method == 'POST':
        nome, uc, preco, uu, fator, qtd_emb, unid_emb, vinc, fornecedor, aproveitamento = _parse_insumo_form(request.form)
        db = get_db()
        db.execute('''INSERT INTO insumos
                      (nome,unidade_compra,preco_compra,unidade_uso,fator_conversao,
                       qtd_por_embalagem,unid_embalagem,produto_vinculado_id,fornecedor,aproveitamento)
                      VALUES(?,?,?,?,?,?,?,?,?,?)''',
                   (nome, uc, preco, uu, fator, qtd_emb, unid_emb, vinc, fornecedor, aproveitamento))
        db.commit(); db.close()
        flash(f'Insumo "{nome}" cadastrado.', 'success')
        return redirect(url_for('insumos_lista'))
    return render_template('insumo_form.html', ins=None, CONV=CONV,
                           produtos=_get_produtos_para_vincular())

@app.route('/insumos/<int:id>/editar', methods=['GET','POST'])
def insumo_editar(id):
    db = get_db()
    ins = db.execute('SELECT * FROM insumos WHERE id=?', (id,)).fetchone()
    if request.method == 'POST':
        nome, uc, preco, uu, fator, qtd_emb, unid_emb, vinc, fornecedor, aproveitamento = _parse_insumo_form(request.form)
        preco_antigo = ins['preco_compra']
        fator_antigo = ins['fator_conversao'] or 1
        db.execute('''UPDATE insumos
                      SET nome=?,unidade_compra=?,preco_compra=?,unidade_uso=?,
                          fator_conversao=?,qtd_por_embalagem=?,unid_embalagem=?,
                          produto_vinculado_id=?,fornecedor=?,aproveitamento=?
                      WHERE id=?''',
                   (nome, uc, preco, uu, fator, qtd_emb, unid_emb, vinc, fornecedor, aproveitamento, id))
        unit_ant = (preco_antigo or 0) / fator_antigo
        unit_novo = (preco or 0) / (fator or 1)
        if round(unit_ant, 6) != round(unit_novo, 6):
            db.execute('INSERT INTO historico_precos (insumo_id, preco_anterior, preco_novo) VALUES (?,?,?)',
                       (id, unit_ant, unit_novo))
            _recalcular_historico_fichas(db, id)
        db.commit(); db.close()
        flash('Insumo atualizado.', 'success')
        return redirect(url_for('insumos_lista'))
    db.close()
    return render_template('insumo_form.html', ins=ins, CONV=CONV,
                           produtos=_get_produtos_para_vincular())

@app.route('/insumos/<int:id>/preco', methods=['POST'])
def insumo_preco(id):
    val = request.json.get('preco', '')
    p = float(val) if str(val).strip() not in ('', 'null', 'None') else None
    db = get_db()
    antigo = db.execute('SELECT preco_compra, fator_conversao FROM insumos WHERE id=?', (id,)).fetchone()
    db.execute('UPDATE insumos SET preco_compra=? WHERE id=?', (p, id))
    if antigo and antigo['preco_compra'] != p:
        fator = antigo['fator_conversao'] or 1
        db.execute('INSERT INTO historico_precos (insumo_id, preco_anterior, preco_novo) VALUES (?,?,?)',
                   (id, (antigo['preco_compra'] or 0) / fator, (p or 0) / fator))
        _recalcular_historico_fichas(db, id)
    db.commit(); db.close()
    return jsonify(ok=True)

@app.route('/insumos/<int:id>/historico')
def insumo_historico(id):
    db = get_db()
    nome = db.execute('SELECT nome FROM insumos WHERE id=?', (id,)).fetchone()
    hist = db.execute('''
        SELECT preco_anterior, preco_novo, data_registro
        FROM historico_precos WHERE insumo_id=? ORDER BY id ASC
    ''', (id,)).fetchall()
    db.close()
    return jsonify(nome=nome['nome'] if nome else '', historico=[dict(r) for r in hist])

@app.route('/insumos/<int:id>/estoque', methods=['POST'])
def insumo_estoque(id):
    val = request.json.get('estoque', 0)
    v = float(val) if str(val).strip() not in ('', 'null', 'None') else 0.0
    db = get_db()
    db.execute('UPDATE insumos SET estoque_atual=? WHERE id=?', (v, id))
    db.commit(); db.close()
    return jsonify(ok=True)

@app.route('/insumos/<int:id>/estoque_central', methods=['POST'])
def insumo_estoque_central(id):
    val = request.json.get('estoque', 0)
    v = float(val) if str(val).strip() not in ('', 'null', 'None') else 0.0
    db = get_db()
    db.execute('UPDATE insumos SET estoque_central=? WHERE id=?', (v, id))
    db.commit(); db.close()
    return jsonify(ok=True)

@app.route('/insumos/<int:id>/estoque_minimo', methods=['POST'])
def insumo_estoque_minimo(id):
    val = request.json.get('estoque_minimo', 0)
    v = float(val) if str(val).strip() not in ('', 'null', 'None') else 0.0
    db = get_db()
    db.execute('UPDATE insumos SET estoque_minimo=? WHERE id=?', (v, id))
    db.commit(); db.close()
    return jsonify(ok=True)

@app.route('/insumos/<int:id>/uso')
def insumo_uso(id):
    db = get_db()
    rows = db.execute('''
        SELECT p.nome, p.categoria, ft.quantidade, i.unidade_uso
        FROM ficha_tecnica ft
        JOIN produtos p ON p.id = ft.produto_id
        JOIN insumos i ON i.id = ft.insumo_id
        WHERE ft.insumo_id = ?
        ORDER BY p.nome
    ''', (id,)).fetchall()
    db.close()
    return jsonify([{'produto': r['nome'], 'categoria': r['categoria'] or '—',
                     'quantidade': r['quantidade'], 'unidade': r['unidade_uso']} for r in rows])

@app.route('/insumos/<int:id>/excluir', methods=['POST'])
def insumo_excluir(id):
    db = get_db()
    uso = db.execute('SELECT COUNT(*) FROM ficha_tecnica WHERE insumo_id=?', (id,)).fetchone()[0]
    if uso:
        db.close()
        flash('Insumo em uso em fichas técnicas — não pode ser excluído.', 'danger')
        return redirect(url_for('insumos_lista'))
    nome = db.execute('SELECT nome FROM insumos WHERE id=?', (id,)).fetchone()['nome']
    db.execute('DELETE FROM insumos WHERE id=?', (id,))
    db.commit(); db.close()
    flash(f'"{nome}" excluído.', 'success')
    return redirect(url_for('insumos_lista'))

@app.route('/insumos/evolucao-precos')
def insumos_evolucao_precos():
    db = get_db()
    records = db.execute('''
        SELECT h.insumo_id, i.nome insumo_nome, i.unidade_compra,
               h.preco_anterior, h.preco_novo,
               date(h.data_registro) data_dia
        FROM historico_precos h
        JOIN insumos i ON i.id = h.insumo_id
        WHERE h.preco_anterior IS NOT NULL
        ORDER BY h.insumo_id, h.id
    ''').fetchall()
    db.close()

    dates = sorted(set(r['data_dia'] for r in records))

    by_id = {}
    for r in records:
        iid = r['insumo_id']
        if iid not in by_id:
            by_id[iid] = {'nome': r['insumo_nome'], 'unidade': r['unidade_compra'] or '',
                          'inicial': r['preco_anterior'], 'por_data': {}}
        by_id[iid]['por_data'][r['data_dia']] = r['preco_novo']

    for iid, d in by_id.items():
        ultimo = d['inicial']
        for dt in dates:
            if dt not in d['por_data']:
                d['por_data'][dt] = ultimo
            else:
                ultimo = d['por_data'][dt]
        d['atual'] = d['por_data'][dates[-1]] if dates else d['inicial']
        d['var_abs'] = d['atual'] - d['inicial']
        d['var_pct'] = (d['var_abs'] / d['inicial'] * 100) if d['inicial'] else 0
        spark = [d['inicial']] + [d['por_data'][dt] for dt in dates]
        d['spark_vals'] = ','.join(str(round(v, 6)) for v in spark)

    insumos = sorted(by_id.values(), key=lambda x: abs(x['var_pct']), reverse=True)
    return render_template('evolucao_insumos.html', insumos=insumos, dates=dates)

# ── Fichas Técnicas ──────────────────────────────────────────
@app.route('/fichas/evolucao-custos')
def fichas_evolucao_custos():
    db = get_db()
    records = db.execute('''
        SELECT hcf.produto_id, p.nome produto_nome, p.categoria,
               hcf.custo_anterior, hcf.custo_novo,
               date(hcf.data_registro) data_dia
        FROM historico_custo_fichas hcf
        JOIN produtos p ON p.id = hcf.produto_id
        WHERE hcf.custo_anterior IS NOT NULL
        ORDER BY hcf.produto_id, hcf.id
    ''').fetchall()
    db.close()

    dates = sorted(set(r['data_dia'] for r in records))

    by_pid = {}
    for r in records:
        pid = r['produto_id']
        if pid not in by_pid:
            by_pid[pid] = {'nome': r['produto_nome'], 'categoria': r['categoria'] or '',
                           'inicial': r['custo_anterior'], 'por_data': {}}
        by_pid[pid]['por_data'][r['data_dia']] = r['custo_novo']

    # Preenche datas ausentes com último custo conhecido
    for pid, d in by_pid.items():
        ultimo = d['inicial']
        for dt in dates:
            if dt not in d['por_data']:
                d['por_data'][dt] = ultimo
            else:
                ultimo = d['por_data'][dt]
        d['atual'] = d['por_data'][dates[-1]] if dates else d['inicial']
        d['var_abs'] = d['atual'] - d['inicial']
        d['var_pct'] = (d['var_abs'] / d['inicial'] * 100) if d['inicial'] else 0
        # Pré-computa valores para sparkline: inicial + custo em cada data
        spark = [d['inicial']] + [d['por_data'][dt] for dt in dates]
        d['spark_vals'] = ','.join(str(round(v, 6)) for v in spark)

    produtos = sorted(by_pid.values(), key=lambda x: abs(x['var_pct']), reverse=True)
    return render_template('evolucao_custos.html', produtos=produtos, dates=dates)

@app.route('/fichas/<int:id>/revisao', methods=['POST'])
def ficha_revisao(id):
    status = request.json.get('status', 'nao_revisada')
    if status not in ('nao_revisada', 'em_revisao', 'revisada'):
        return jsonify(ok=False), 400
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    data_revisao = now if status == 'revisada' else None
    db = get_db()
    db.execute('UPDATE produtos SET status_revisao=?, data_revisao=? WHERE id=?', (status, data_revisao, id))
    db.commit(); db.close()
    return jsonify(ok=True, status=status, data_revisao=data_revisao)

@app.route('/fichas/<int:id>/revisao-custo', methods=['POST'])
def ficha_revisao_custo(id):
    status = request.json.get('status', 'nao_revisado')
    if status not in ('nao_revisado', 'em_revisao', 'revisado'):
        return jsonify(ok=False), 400
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    data_revisao_custo = now if status == 'revisado' else None
    db = get_db()
    db.execute('UPDATE produtos SET status_revisao_custo=?, data_revisao_custo=? WHERE id=?', (status, data_revisao_custo, id))
    db.commit(); db.close()
    return jsonify(ok=True, status=status, data_revisao_custo=data_revisao_custo)

@app.route('/insumos/<int:id>/revisao-custo', methods=['POST'])
def insumo_revisao_custo(id):
    status = request.json.get('status', 'nao_revisado')
    if status not in ('nao_revisado', 'em_revisao', 'revisado'):
        return jsonify(ok=False), 400
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    data_revisao_custo = now if status == 'revisado' else None
    db = get_db()
    db.execute('UPDATE insumos SET status_revisao_custo=?, data_revisao_custo=? WHERE id=?', (status, data_revisao_custo, id))
    db.commit(); db.close()
    return jsonify(ok=True, status=status, data_revisao_custo=data_revisao_custo)

@app.route('/insumos/<int:id>/observacao', methods=['POST'])
def insumo_observacao(id):
    obs = request.json.get('observacao', '').strip()
    db = get_db()
    db.execute('UPDATE insumos SET observacao=? WHERE id=?', (obs or None, id))
    db.commit(); db.close()
    return jsonify(ok=True)

@app.route('/fichas/<int:id>/observacao', methods=['POST'])
def ficha_observacao(id):
    obs = request.json.get('observacao', '').strip()
    db = get_db()
    db.execute('UPDATE produtos SET observacao=? WHERE id=?', (obs or None, id))
    db.commit(); db.close()
    return jsonify(ok=True)

@app.route('/insumos/<int:id>/revisao-nome', methods=['POST'])
def insumo_revisao_nome(id):
    status = request.json.get('status', 'nao_revisado')
    if status not in ('nao_revisado', 'em_revisao', 'revisado'):
        return jsonify(ok=False), 400
    db = get_db()
    db.execute('UPDATE insumos SET status_revisao_nome=? WHERE id=?', (status, id))
    db.commit(); db.close()
    return jsonify(ok=True, status=status)

@app.route('/fichas/<int:id>/historico-custo')
def ficha_historico_custo(id):
    db = get_db()
    nome = db.execute('SELECT nome FROM produtos WHERE id=?', (id,)).fetchone()
    hist = db.execute('SELECT custo_anterior, custo_novo, data_registro FROM historico_custo_fichas WHERE produto_id=? ORDER BY id ASC', (id,)).fetchall()
    db.close()
    return jsonify(nome=nome['nome'] if nome else '', historico=[dict(r) for r in hist])

@app.route('/fichas')
def fichas_lista():
    db = get_db()
    prods = db.execute('SELECT * FROM produtos ORDER BY nome').fetchall()
    hist_rows = db.execute('''
        SELECT h.produto_id, h.custo_anterior, h.custo_novo, h.data_registro
        FROM historico_custo_fichas h
        WHERE h.id = (SELECT MAX(id) FROM historico_custo_fichas h2 WHERE h2.produto_id = h.produto_id)
    ''').fetchall()
    mudancas_custo = {r['produto_id']: dict(r) for r in hist_rows}
    # Custo revisão: calculado a partir do status dos insumos de cada ficha
    rev_rows = db.execute('''
        SELECT ft.produto_id,
               SUM(CASE WHEN COALESCE(i.status_revisao_custo,'nao_revisado')='revisado' THEN 1 ELSE 0 END) AS revisados,
               COUNT(*) AS total
        FROM ficha_tecnica ft JOIN insumos i ON i.id = ft.insumo_id
        GROUP BY ft.produto_id
    ''').fetchall()
    custo_revisao_map = {}
    for rr in rev_rows:
        if rr['total'] == 0 or rr['revisados'] == 0:
            custo_revisao_map[rr['produto_id']] = 'nao_revisado'
        elif rr['revisados'] == rr['total']:
            custo_revisao_map[rr['produto_id']] = 'revisado'
        else:
            custo_revisao_map[rr['produto_id']] = 'incompleta'
    db.close()
    resultado = [{'p': p, **dict(zip(['custo','completo'], custo_produto(p['id'])))} for p in prods]
    for r in resultado:
        r['custo_revisao'] = custo_revisao_map.get(r['p']['id'], 'nao_revisado')
    cat_map = {nome: cls for nome, cls in CATEGORIAS}
    total_custo = sum(r['custo'] for r in resultado if r['completo'])
    n_completos = sum(1 for r in resultado if r['completo'])
    return render_template('fichas_lista.html', resultado=resultado, categorias=CATEGORIAS, cat_map=cat_map,
                           total_custo=total_custo, n_completos=n_completos, mudancas_custo=mudancas_custo)

@app.route('/fichas/<int:id>/toggle-ativo', methods=['POST'])
def ficha_toggle_ativo(id):
    db = get_db()
    prod = db.execute('SELECT ativo FROM produtos WHERE id=?', (id,)).fetchone()
    novo = 0 if (prod['ativo'] if prod['ativo'] is not None else 1) else 1
    db.execute('UPDATE produtos SET ativo=? WHERE id=?', (novo, id))
    db.commit()
    db.close()
    from flask import jsonify
    return jsonify({'ativo': novo})

@app.route('/fichas/<int:id>/categoria', methods=['POST'])
def ficha_set_categoria(id):
    cat = request.json.get('categoria', '').strip() or None
    db = get_db()
    db.execute('UPDATE produtos SET categoria=? WHERE id=?', (cat, id))
    db.commit(); db.close()
    return jsonify({'ok': True, 'categoria': cat})

@app.route('/fichas/<int:id>')
def ficha_ver(id):
    db = get_db()
    prod = db.execute('SELECT * FROM produtos WHERE id=?', (id,)).fetchone()
    if not prod:
        db.close(); return redirect(url_for('fichas_lista'))
    items = db.execute('''
        SELECT ft.id ft_id, ft.quantidade, i.id ins_id, i.nome ins_nome,
               i.unidade_uso, i.preco_compra, i.fator_conversao,
               i.produto_vinculado_id, p2.nome as nome_vinc, p2.rendimento as rend_vinc,
               i.status_revisao_custo, i.data_revisao_custo,
               COALESCE(ft.aproveitamento, 100) AS aproveitamento
        FROM ficha_tecnica ft
        JOIN insumos i ON i.id=ft.insumo_id
        LEFT JOIN produtos p2 ON p2.id = i.produto_vinculado_id
        WHERE ft.produto_id=? ORDER BY i.nome''', (id,)).fetchall()
    db.close()
    total, incompleto = 0.0, False
    itens = []
    for it in items:
        c = None
        origem = 'preco'
        aprov = (it['aproveitamento'] or 100) / 100
        qtd_compra = it['quantidade'] / aprov
        if it['produto_vinculado_id']:
            sub, sub_ok = custo_produto(it['produto_vinculado_id'])
            if sub_ok:
                rend = it['rend_vinc'] or 1
                c = qtd_compra * (sub / rend) / it['fator_conversao']
                total += c
                origem = 'ficha'
            else:
                incompleto = True
        elif it['preco_compra'] is not None:
            c = qtd_compra * it['preco_compra'] / it['fator_conversao']
            total += c
        else:
            incompleto = True
        itens.append({**dict(it), 'custo': c, 'origem': origem, 'qtd_compra': qtd_compra})
    cat_map = {nome: cls for nome, cls in CATEGORIAS}
    return render_template('ficha_ver.html', prod=prod, itens=itens, total=total, incompleto=incompleto, cat_map=cat_map)

@app.route('/fichas/novo', methods=['GET','POST'])
def ficha_nova():
    if request.method == 'POST':
        d = request.form
        db = get_db()
        vref = d.get('valor_referencia','').strip()
        vref = float(vref) if vref else None
        db.execute('INSERT INTO produtos (nome,descricao,rendimento,unidade_rendimento,categoria,valor_referencia) VALUES(?,?,?,?,?,?)',
                   (d['nome'].strip(), d.get('descricao','').strip(),
                    float(d.get('rendimento') or 1), d.get('unidade_rendimento','un'),
                    d.get('categoria','').strip() or None, vref))
        pid = db.execute('SELECT last_insert_rowid()').fetchone()[0]
        db.commit(); db.close()
        flash('Produto criado. Adicione os ingredientes.', 'success')
        return redirect(url_for('ficha_editar', id=pid))
    return render_template('ficha_form_produto.html', prod=None, categorias=CATEGORIAS)

@app.route('/fichas/<int:id>/editar', methods=['GET','POST'])
def ficha_editar(id):
    db = get_db()
    prod = db.execute('SELECT * FROM produtos WHERE id=?', (id,)).fetchone()
    if not prod:
        db.close(); return redirect(url_for('fichas_lista'))
    if request.method == 'POST':
        d = request.form
        vref = d.get('valor_referencia','').strip()
        vref = float(vref) if vref else None
        novo_nome = d['nome'].strip()
        db.execute('UPDATE produtos SET nome=?,descricao=?,rendimento=?,unidade_rendimento=?,categoria=?,valor_referencia=? WHERE id=?',
                   (novo_nome, d.get('descricao','').strip(),
                    float(d.get('rendimento') or 1), d.get('unidade_rendimento','un'),
                    d.get('categoria','').strip() or None, vref, id))
        # Sincroniza o nome do insumo vinculado a este produto
        db.execute('UPDATE insumos SET nome=? WHERE produto_vinculado_id=?', (novo_nome, id))
        db.commit()
        flash('Produto atualizado.', 'success')
        prod = db.execute('SELECT * FROM produtos WHERE id=?', (id,)).fetchone()
    items = db.execute('''
        SELECT ft.id ft_id, ft.quantidade, ft.aproveitamento,
               i.id ins_id, i.nome ins_nome, i.unidade_uso,
               i.produto_vinculado_id, i.preco_compra, i.fator_conversao,
               p2.rendimento AS rend_vinc
        FROM ficha_tecnica ft JOIN insumos i ON i.id=ft.insumo_id
        LEFT JOIN produtos p2 ON p2.id = i.produto_vinculado_id
        WHERE ft.produto_id=? ORDER BY i.nome''', (id,)).fetchall()
    todos = db.execute('SELECT id, nome, unidade_uso FROM insumos ORDER BY nome').fetchall()
    produtos_lista = db.execute('SELECT id, nome, unidade_rendimento FROM produtos WHERE id != ? ORDER BY nome', (id,)).fetchall()
    db.close()
    # Custo por ingrediente
    custos_items = []
    for it in items:
        aprov = (it['aproveitamento'] or 100) / 100
        qtd_compra = it['quantidade'] / aprov
        if it['produto_vinculado_id']:
            sub, ok = custo_produto(it['produto_vinculado_id'])
            rend = it['rend_vinc'] or 1
            c = qtd_compra * (sub / rend) / (it['fator_conversao'] or 1) if ok else None
        elif it['preco_compra'] is not None:
            c = qtd_compra * it['preco_compra'] / (it['fator_conversao'] or 1)
        else:
            c = None
        custos_items.append(c)
    custo_total, custo_completo = custo_produto(id)
    return render_template('ficha_editar.html', prod=prod, items=items, todos=todos,
                           categorias=CATEGORIAS, custos_items=custos_items,
                           custo_total=custo_total, custo_completo=custo_completo,
                           produtos_lista=produtos_lista)

@app.route('/fichas/<int:id>/ingrediente/add', methods=['POST'])
def ficha_add_ing(id):
    ins_id = int(request.form['insumo_id'])
    qty    = float(request.form['quantidade'])
    db = get_db()
    ex = db.execute('SELECT id FROM ficha_tecnica WHERE produto_id=? AND insumo_id=?', (id, ins_id)).fetchone()
    if ex:
        db.execute('UPDATE ficha_tecnica SET quantidade=? WHERE id=?', (qty, ex['id']))
    else:
        db.execute('INSERT INTO ficha_tecnica (produto_id,insumo_id,quantidade) VALUES(?,?,?)', (id, ins_id, qty))
    db.commit(); db.close()
    flash('Ingrediente adicionado.', 'success')
    return redirect(url_for('ficha_editar', id=id))

@app.route('/fichas/<int:id>/subreceita/add', methods=['POST'])
def ficha_add_subreceita(id):
    sub_prod_id = int(request.form['produto_id'])
    qty = float(request.form['quantidade'])
    db = get_db()
    sub_prod = db.execute('SELECT id, nome, unidade_rendimento FROM produtos WHERE id=?', (sub_prod_id,)).fetchone()
    if not sub_prod:
        db.close()
        flash('Sub-receita não encontrada.', 'danger')
        return redirect(url_for('ficha_editar', id=id))
    # Find or create the linking insumo
    ins = db.execute('SELECT id FROM insumos WHERE produto_vinculado_id=?', (sub_prod_id,)).fetchone()
    if ins:
        ins_id = ins['id']
    else:
        uu = sub_prod['unidade_rendimento'] or 'un'
        db.execute('''INSERT INTO insumos (nome, unidade_compra, preco_compra, unidade_uso,
                      fator_conversao, produto_vinculado_id)
                      VALUES(?,?,NULL,?,1,?)''', (sub_prod['nome'], uu, uu, sub_prod_id))
        ins_id = db.execute('SELECT last_insert_rowid()').fetchone()[0]
    ex = db.execute('SELECT id FROM ficha_tecnica WHERE produto_id=? AND insumo_id=?', (id, ins_id)).fetchone()
    if ex:
        db.execute('UPDATE ficha_tecnica SET quantidade=? WHERE id=?', (qty, ex['id']))
    else:
        db.execute('INSERT INTO ficha_tecnica (produto_id,insumo_id,quantidade) VALUES(?,?,?)', (id, ins_id, qty))
    db.commit(); db.close()
    flash(f'Sub-receita "{sub_prod["nome"]}" adicionada.', 'success')
    return redirect(url_for('ficha_editar', id=id))

@app.route('/fichas/<int:pid>/ingrediente/<int:ft_id>/ajustar', methods=['POST'])
def ficha_ajustar_ing(pid, ft_id):
    try:
        qty = float(request.form['quantidade'])
        aprov_str = request.form.get('aproveitamento', '').strip()
        aprov = float(aprov_str) if aprov_str else None
        if aprov is not None:
            aprov = max(1.0, min(100.0, aprov))
        if qty > 0:
            db = get_db()
            if aprov is not None:
                db.execute('UPDATE ficha_tecnica SET quantidade=?, aproveitamento=? WHERE id=? AND produto_id=?',
                           (qty, aprov, ft_id, pid))
            else:
                db.execute('UPDATE ficha_tecnica SET quantidade=? WHERE id=? AND produto_id=?', (qty, ft_id, pid))
            db.commit(); db.close()
    except Exception:
        pass
    return redirect(url_for('ficha_editar', id=pid))

@app.route('/fichas/<int:pid>/ingrediente/<int:ft_id>/aproveitamento', methods=['POST'])
def ficha_aprov_ing(pid, ft_id):
    from flask import jsonify
    try:
        aprov = float(request.json.get('aproveitamento', 100))
        aprov = max(1.0, min(100.0, aprov))
        db = get_db()
        db.execute('UPDATE ficha_tecnica SET aproveitamento=? WHERE id=? AND produto_id=?', (aprov, ft_id, pid))
        db.commit(); db.close()
        return jsonify(ok=True, aproveitamento=aprov)
    except Exception as e:
        return jsonify(ok=False, erro=str(e)), 400

@app.route('/fichas/<int:pid>/ingrediente/<int:ft_id>/remover', methods=['POST'])
def ficha_rem_ing(pid, ft_id):
    db = get_db()
    db.execute('DELETE FROM ficha_tecnica WHERE id=? AND produto_id=?', (ft_id, pid))
    db.commit(); db.close()
    return redirect(url_for('ficha_editar', id=pid))

@app.route('/fichas/<int:id>/excluir', methods=['POST'])
def ficha_excluir(id):
    db = get_db()
    nome = db.execute('SELECT nome FROM produtos WHERE id=?', (id,)).fetchone()['nome']
    db.execute('DELETE FROM produtos WHERE id=?', (id,))
    db.commit(); db.close()
    flash(f'"{nome}" excluído.', 'success')
    return redirect(url_for('fichas_lista'))

# ── Programação de Produção ──────────────────────────────────
@app.route('/producao', methods=['GET','POST'])
def producao():
    db = get_db()
    prods = db.execute('SELECT * FROM produtos ORDER BY nome').fetchall()
    db.close()
    custos = {p['id']: dict(zip(['custo','completo'], custo_produto(p['id']))) for p in prods}

    resultado = None
    qtds = {p['id']: 0 for p in prods}

    if request.method == 'POST':
        for p in prods:
            try:
                qtds[p['id']] = float(request.form.get(f'qty_{p["id"]}', 0) or 0)
            except:
                qtds[p['id']] = 0

        db = get_db()
        ins_map = {}

        for p in prods:
            qty = qtds[p['id']]
            if qty <= 0:
                continue
            rend = p['rendimento'] or 1
            _expand_ingredientes(db, p['id'], qty / rend, ins_map)
        db.close()

        lista = []
        _inteiras_prod = {'un','und','unid','pct','pacote','cx','caixa','fardo','saco','bd','bandeja',
                          'dz','duzia','rl','rolo','lt','lata','sc','pt','pote','balde','galão','galao','garrafa','frasco','maço','maco','mc','bisnaga','bisn'}
        for iid, d in sorted(ins_map.items(), key=lambda x: x[1]['nome']):
            estoque = d.get('estoque_atual') or 0.0
            qtd_necessaria = max(0.0, d['qtd_uso'] - estoque)
            _qtd = qtd_necessaria / d['fator_conversao']
            if (d['unidade_compra'] or '').lower() in _inteiras_prod:
                _qtd = math.ceil(_qtd)
            valor_compra = (_qtd * d['preco_compra']) if d['preco_compra'] is not None else None
            sobra = _qtd * d['fator_conversao'] - qtd_necessaria
            sobra_valor = (sobra / d['fator_conversao'] * d['preco_compra']) if (sobra > 0 and d['preco_compra'] is not None) else None
            lista.append({**d, 'qtd_necessaria': d['qtd_uso'], 'estoque_disponivel': estoque,
                          'qtd_compra': _qtd, 'valor': valor_compra, 'sobra': sobra, 'sobra_valor': sobra_valor})

        custo_total = sum(i['valor'] for i in lista if i['valor'] is not None)
        sobra_total = sum(i['sobra_valor'] for i in lista if i.get('sobra_valor') is not None)
        resultado = {'lista': lista, 'custo_total': custo_total, 'sobra_total': sobra_total,
                     'falta_preco': any(i['preco_compra'] is None for i in lista)}

    cat_map = {nome: cls for nome, cls in CATEGORIAS}
    return render_template('producao.html', prods=prods, qtds=qtds, resultado=resultado, custos=custos, categorias=CATEGORIAS, cat_map=cat_map)


def _expand_ingredientes(db, produto_id, scale, ins_map, _visited=None):
    """Expande recursivamente sub-receitas até os insumos base."""
    if _visited is None:
        _visited = set()
    if produto_id in _visited:
        return
    _visited = _visited | {produto_id}

    rows = db.execute('''
        SELECT ft.quantidade, i.id ins_id, i.nome, i.unidade_uso,
               i.unidade_compra, i.preco_compra, i.fator_conversao,
               i.produto_vinculado_id,
               COALESCE(i.estoque_atual, 0) as estoque_atual,
               p2.rendimento AS rend_vinc
        FROM ficha_tecnica ft JOIN insumos i ON i.id=ft.insumo_id
        LEFT JOIN produtos p2 ON p2.id = i.produto_vinculado_id
        WHERE ft.produto_id=?''', (produto_id,)).fetchall()

    for r in rows:
        if r['produto_vinculado_id']:
            # Sub-receita: calcula escala e expande recursivamente
            rend = r['rend_vinc'] or 1
            fator = r['fator_conversao'] or 1
            sub_scale = scale * r['quantidade'] / fator / rend
            _expand_ingredientes(db, r['produto_vinculado_id'], sub_scale, ins_map, _visited)
        else:
            iid = r['ins_id']
            qtd_uso = r['quantidade'] * scale
            if iid not in ins_map:
                ins_map[iid] = {'nome': r['nome'], 'unidade_uso': r['unidade_uso'],
                                'unidade_compra': r['unidade_compra'],
                                'preco_compra': r['preco_compra'],
                                'fator_conversao': r['fator_conversao'] or 1,
                                'estoque_atual': r['estoque_atual'],
                                'qtd_uso': 0.0, 'valor': 0.0}
            ins_map[iid]['qtd_uso'] += qtd_uso
            if r['preco_compra'] is not None:
                ins_map[iid]['valor'] += qtd_uso * r['preco_compra'] / (r['fator_conversao'] or 1)


def _calcular_lista(prods, form, estoque_override=None):
    db = get_db()
    ins_map = {}
    for p in prods:
        try:
            qty = float(form.get(f'qty_{p["id"]}', 0) or 0)
        except Exception:
            qty = 0
        if qty <= 0:
            continue
        rend = p['rendimento'] or 1
        _expand_ingredientes(db, p['id'], qty / rend, ins_map)
    db.close()

    lista = []
    _inteiras = {'un','und','unid','pct','pacote','cx','caixa','fardo','saco','bd','bandeja',
                 'dz','duzia','rl','rolo','lt','lata','sc','pt','pote','fatia','balde','galão','galao','garrafa','frasco','maço','maco','mc','bisnaga','bisn'}
    for iid, d in sorted(ins_map.items(), key=lambda x: x[1]['nome']):
        if estoque_override is not None:
            estoque = float(estoque_override.get(iid, 0.0))
        else:
            estoque = d.get('estoque_atual') or 0.0
        qtd_nec = max(0.0, d['qtd_uso'] - estoque)
        _qtd = qtd_nec / d['fator_conversao']
        if (d['unidade_compra'] or '').lower() in _inteiras:
            _qtd = math.ceil(_qtd)
        valor_compra = (_qtd * d['preco_compra']) if d['preco_compra'] is not None else None
        sobra = _qtd * d['fator_conversao'] - qtd_nec
        sobra_valor = (sobra / d['fator_conversao'] * d['preco_compra']) if (sobra > 0 and d['preco_compra'] is not None) else None
        lista.append({**d, 'id': iid, 'qtd_necessaria': d['qtd_uso'], 'estoque_disponivel': estoque,
                      'qtd_compra': _qtd, 'valor': valor_compra, 'sobra': sobra, 'sobra_valor': sobra_valor})
    custo_total = sum(i['valor'] for i in lista if i['valor'] is not None)
    return lista, custo_total


@app.route('/producao/exportar', methods=['POST'])
def producao_exportar():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash('openpyxl não instalado. Execute: pip install openpyxl', 'danger')
        return redirect(url_for('producao'))

    db = get_db()
    prods = db.execute('SELECT * FROM produtos ORDER BY nome').fetchall()
    db.close()
    lista, custo_total = _calcular_lista(prods, request.form)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Lista de Compras'

    # Cores
    verde_escuro = '375623'
    verde_claro  = 'E8F5E9'
    laranja      = 'FFF3E0'
    cinza_header = '1A3A5C'
    branco       = 'FFFFFF'

    def cell_fill(color):
        return PatternFill('solid', fgColor=color)

    def border_thin():
        s = Side(style='thin', color='CCCCCC')
        return Border(left=s, right=s, top=s, bottom=s)

    # Título
    ws.merge_cells('A1:I1')
    ws['A1'] = f'Lista de Compras — Mercado Selectus'
    ws['A1'].font = Font(bold=True, size=14, color=branco)
    ws['A1'].fill = cell_fill(cinza_header)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:I2')
    ws['A2'] = f'Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font = Font(italic=True, size=10, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 16

    # Cabeçalho
    headers = ['#', 'Insumo', 'Qtd. Necessária', 'Unid. Uso', 'Em Estoque',
               'Qtd. a Comprar', 'Unid. Compra', 'Preço / Unid. (R$)', 'Valor Previsto (R$)']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = Font(bold=True, color=branco, size=10)
        c.fill = cell_fill(verde_escuro)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border_thin()
    ws.row_dimensions[4].height = 32

    # Dados
    for i, item in enumerate(lista, 1):
        row = i + 4
        fill = cell_fill(verde_claro) if i % 2 == 0 else cell_fill(branco)
        if item['qtd_compra'] == 0:
            fill = cell_fill('F1F8E9')

        vals = [
            i,
            item['nome'],
            round(item['qtd_necessaria'], 2),
            item['unidade_uso'],
            round(item['estoque_disponivel'], 2) if item['estoque_disponivel'] else '—',
            item['qtd_compra'] if item['qtd_compra'] > 0 else 'Coberto',
            item['unidade_compra'],
            item['preco_compra'] if item['preco_compra'] else '—',
            round(item['valor'], 2) if item['preco_compra'] and item['qtd_compra'] > 0 else '—',
        ]
        aligns = ['center','left','right','center','right','right','center','right','right']
        for col, (val, aln) in enumerate(zip(vals, aligns), 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = fill
            c.alignment = Alignment(horizontal=aln, vertical='center')
            c.border = border_thin()
            c.font = Font(size=10)
            if col in (3, 5, 6) and isinstance(val, (int, float)):
                c.number_format = '#,##0.00'
            if col in (8, 9) and isinstance(val, float):
                c.number_format = 'R$ #,##0.0000'

    # Rodapé total
    total_row = len(lista) + 5
    ws.merge_cells(f'A{total_row}:H{total_row}')
    ws[f'A{total_row}'] = 'VALOR TOTAL PREVISTO DE COMPRAS'
    ws[f'A{total_row}'].font = Font(bold=True, color=branco, size=11)
    ws[f'A{total_row}'].fill = cell_fill(verde_escuro)
    ws[f'A{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'I{total_row}'] = round(custo_total, 2)
    ws[f'I{total_row}'].font = Font(bold=True, color=branco, size=11)
    ws[f'I{total_row}'].fill = cell_fill(verde_escuro)
    ws[f'I{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'I{total_row}'].number_format = 'R$ #,##0.00'
    ws.row_dimensions[total_row].height = 24

    # Larguras das colunas
    widths = [5, 40, 16, 12, 14, 16, 14, 20, 20]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'lista_compras_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── Requisições de Produção ────────────────────────────────────────────────

@app.route('/requisicoes')
def requisicoes_lista():
    db = get_db()
    reqs = db.execute('''
        SELECT r.id, r.data, r.status, r.observacao, r.criado_em, r.fechado_em,
               COUNT(ri.id) as n_itens
        FROM requisicoes r
        LEFT JOIN requisicao_itens ri ON ri.requisicao_id = r.id
        GROUP BY r.id
        ORDER BY r.data DESC, r.id DESC
    ''').fetchall()
    db.close()
    return render_template('requisicoes_lista.html', reqs=reqs)


@app.route('/requisicoes/nova', methods=['GET','POST'])
def requisicao_nova():
    db = get_db()
    if request.method == 'POST':
        data      = request.form.get('data') or date.today().isoformat()
        obs       = request.form.get('observacao', '').strip()
        insumo_ids = request.form.getlist('insumo_id')
        qtds       = request.form.getlist('qtd_enviada')

        itens = []
        for iid, q in zip(insumo_ids, qtds):
            try:
                qtd = float(q.replace(',', '.'))
                if qtd > 0:
                    itens.append((int(iid), qtd))
            except (ValueError, TypeError):
                pass

        if not itens:
            flash('Adicione ao menos um insumo com quantidade válida.', 'warning')
            insumos = db.execute('SELECT id, nome, unidade_uso, estoque_atual FROM insumos ORDER BY nome COLLATE NOCASE').fetchall()
            db.close()
            return render_template('requisicao_nova.html', insumos=insumos, today=date.today().isoformat())

        db.execute('INSERT INTO requisicoes (data, status, observacao) VALUES (?,?,?)',
                   (data, 'aberto', obs or None))
        req_id = db.execute('SELECT last_insert_rowid()').fetchone()[0]

        for iid, qtd in itens:
            db.execute('INSERT INTO requisicao_itens (requisicao_id, insumo_id, qtd_enviada) VALUES (?,?,?)',
                       (req_id, iid, qtd))
            db.execute('UPDATE insumos SET estoque_atual = COALESCE(estoque_atual,0) - ? WHERE id = ?',
                       (qtd, iid))

        db.commit()
        db.close()
        flash(f'Requisição #{req_id} criada — {len(itens)} insumos enviados para produção.', 'success')
        return redirect(url_for('requisicao_ver', id=req_id))

    insumos = db.execute('SELECT id, nome, unidade_uso, estoque_atual FROM insumos ORDER BY nome COLLATE NOCASE').fetchall()
    db.close()
    return render_template('requisicao_nova.html', insumos=insumos, today=date.today().isoformat())


@app.route('/requisicoes/<int:id>')
def requisicao_ver(id):
    db = get_db()
    req = db.execute('SELECT * FROM requisicoes WHERE id=?', (id,)).fetchone()
    if not req:
        flash('Requisição não encontrada.', 'danger')
        db.close()
        return redirect(url_for('requisicoes_lista'))
    itens = db.execute('''
        SELECT ri.id, ri.qtd_enviada, ri.qtd_retornada,
               i.id as insumo_id, i.nome, i.unidade_uso
        FROM requisicao_itens ri
        JOIN insumos i ON i.id = ri.insumo_id
        WHERE ri.requisicao_id = ?
        ORDER BY i.nome COLLATE NOCASE
    ''', (id,)).fetchall()
    db.close()
    return render_template('requisicao_ver.html', req=req, itens=itens)


@app.route('/requisicoes/<int:id>/fechar', methods=['POST'])
def requisicao_fechar(id):
    db = get_db()
    req = db.execute('SELECT * FROM requisicoes WHERE id=?', (id,)).fetchone()
    if not req or req['status'] != 'aberto':
        flash('Requisição não encontrada ou já fechada.', 'warning')
        db.close()
        return redirect(url_for('requisicoes_lista'))

    item_ids  = request.form.getlist('item_id')
    retornados = request.form.getlist('qtd_retornada')

    for iid, ret in zip(item_ids, retornados):
        try:
            qtd_ret = float(ret.replace(',', '.')) if ret.strip() else 0.0
        except ValueError:
            qtd_ret = 0.0
        item = db.execute('SELECT * FROM requisicao_itens WHERE id=?', (int(iid),)).fetchone()
        if not item:
            continue
        qtd_ret = min(qtd_ret, item['qtd_enviada'])
        db.execute('UPDATE requisicao_itens SET qtd_retornada=? WHERE id=?', (qtd_ret, int(iid)))
        if qtd_ret > 0:
            db.execute('UPDATE insumos SET estoque_atual = COALESCE(estoque_atual,0) + ? WHERE id=?',
                       (qtd_ret, item['insumo_id']))

    db.execute("UPDATE requisicoes SET status='fechado', fechado_em=datetime('now','localtime') WHERE id=?", (id,))
    db.commit()
    db.close()
    flash(f'Requisição #{id} fechada — estoque atualizado com o retorno.', 'success')
    return redirect(url_for('requisicao_ver', id=id))


@app.route('/requisicoes/<int:id>/cancelar', methods=['POST'])
def requisicao_cancelar(id):
    db = get_db()
    req = db.execute('SELECT * FROM requisicoes WHERE id=?', (id,)).fetchone()
    if not req or req['status'] != 'aberto':
        db.close()
        return redirect(url_for('requisicoes_lista'))
    itens = db.execute('SELECT insumo_id, qtd_enviada FROM requisicao_itens WHERE requisicao_id=?', (id,)).fetchall()
    for it in itens:
        db.execute('UPDATE insumos SET estoque_atual = COALESCE(estoque_atual,0) + ? WHERE id=?',
                   (it['qtd_enviada'], it['insumo_id']))
    db.execute("UPDATE requisicoes SET status='cancelado', fechado_em=datetime('now','localtime') WHERE id=?", (id,))
    db.commit()
    db.close()
    flash(f'Requisição #{id} cancelada — estoque restaurado.', 'warning')
    return redirect(url_for('requisicoes_lista'))


@app.route('/requisicoes/<int:id>/excluir', methods=['POST'])
def requisicao_excluir(id):
    db = get_db()
    req = db.execute('SELECT * FROM requisicoes WHERE id=?', (id,)).fetchone()
    if not req:
        db.close()
        return redirect(url_for('requisicoes_lista'))

    itens = db.execute('SELECT insumo_id, qtd_enviada, qtd_retornada FROM requisicao_itens WHERE requisicao_id=?', (id,)).fetchall()

    if req['status'] == 'aberto':
        # Devolve tudo ao estoque
        for it in itens:
            db.execute('UPDATE insumos SET estoque_atual = COALESCE(estoque_atual,0) + ? WHERE id=?',
                       (it['qtd_enviada'], it['insumo_id']))
    elif req['status'] == 'fechado':
        # Devolve apenas o consumo (enviado − retornado), pois retorno já voltou no fechamento
        for it in itens:
            consumo = it['qtd_enviada'] - (it['qtd_retornada'] or 0)
            if consumo > 0:
                db.execute('UPDATE insumos SET estoque_atual = COALESCE(estoque_atual,0) + ? WHERE id=?',
                           (consumo, it['insumo_id']))
    # cancelado: sem efeito no estoque

    db.execute('DELETE FROM requisicoes WHERE id=?', (id,))
    db.commit()
    db.close()
    flash(f'Requisição #{id} excluída — estoque restaurado.', 'success')
    return redirect(url_for('requisicoes_lista'))


@app.route('/requisicoes/consulta')
def requisicao_consulta():
    return render_template('requisicao_consulta.html')


@app.route('/requisicoes/consulta/buscar')
def requisicao_consulta_buscar():
    q = request.args.get('q', '').strip()
    if len(q) < 2:
        return jsonify([])
    db = get_db()
    insumos = db.execute('''
        SELECT id, nome, unidade_uso, unidade_compra, fator_conversao,
               COALESCE(estoque_atual, 0) as estoque_atual,
               COALESCE(estoque_central, 0) as estoque_central
        FROM insumos
        WHERE nome LIKE ?
        ORDER BY nome COLLATE NOCASE
        LIMIT 8
    ''', (f'%{q}%',)).fetchall()

    result = []
    for ins in insumos:
        iid  = ins['id']
        fator = ins['fator_conversao'] or 1
        reqs = db.execute('''
            SELECT ri.id as item_id, r.id as req_id, r.data as req_data, r.observacao as req_obs,
                   ri.qtd_enviada,
                   COALESCE((SELECT SUM(rc.qtd_consumida)
                              FROM requisicao_consumos rc WHERE rc.item_id = ri.id), 0) as total_consumido
            FROM requisicao_itens ri
            JOIN requisicoes r ON r.id = ri.requisicao_id
            WHERE ri.insumo_id = ? AND r.status = 'aberto'
            ORDER BY r.data DESC, r.id DESC
        ''', (iid,)).fetchall()

        reqs_list, req_saldo_total = [], 0.0
        for rq in reqs:
            saldo = max(0.0, rq['qtd_enviada'] - rq['total_consumido'])
            req_saldo_total += saldo
            reqs_list.append({
                'item_id': rq['item_id'], 'req_id': rq['req_id'],
                'req_data': rq['req_data'], 'req_obs': rq['req_obs'],
                'qtd_enviada': rq['qtd_enviada'],
                'total_consumido': rq['total_consumido'], 'saldo': saldo,
            })

        movs = db.execute('''
            SELECT 'transferencia' as tipo, r.data as data,
                   ri.qtd_enviada as qtd, r.id as req_id, ri.id as row_id,
                   NULL as obs, NULL as coz_antes, NULL as coz_depois
            FROM requisicao_itens ri
            JOIN requisicoes r ON r.id = ri.requisicao_id
            WHERE ri.insumo_id = ?
            UNION ALL
            SELECT 'consumo' as tipo, rc.data as data,
                   rc.qtd_consumida as qtd, rc.requisicao_id as req_id, rc.id as row_id,
                   rc.obs as obs, NULL as coz_antes, NULL as coz_depois
            FROM requisicao_consumos rc
            WHERE rc.insumo_id = ?
            UNION ALL
            SELECT 'ajuste' as tipo, al.data as data,
                   NULL as qtd, NULL as req_id, al.id as row_id,
                   al.obs as obs,
                   al.estoque_cozinha_antes as coz_antes,
                   al.estoque_cozinha_depois as coz_depois
            FROM insumo_ajuste_log al
            WHERE al.insumo_id = ?
            ORDER BY data DESC, row_id DESC
            LIMIT 15
        ''', (iid, iid, iid)).fetchall()

        result.append({
            'insumo_id': iid, 'insumo_nome': ins['nome'],
            'unidade_uso': ins['unidade_uso'], 'unidade_compra': ins['unidade_compra'],
            'fator': fator,
            'estoque_atual': ins['estoque_atual'],
            'estoque_central': ins['estoque_central'],
            'est_total': ins['estoque_atual'] + ins['estoque_central'],
            'req_saldo_total': req_saldo_total,
            'requisicoes': reqs_list,
            'movimentacoes': [dict(m) for m in movs],
        })

    db.close()
    return jsonify(result)


@app.route('/requisicoes/<int:id>/consumo', methods=['POST'])
def requisicao_consumo_add(id):
    db = get_db()
    item_id   = int(request.form.get('item_id', 0))
    data      = request.form.get('data') or date.today().isoformat()
    qtd_sobra = float((request.form.get('qtd_sobra', '0') or '0').replace(',', '.'))
    total_cozinha_str = request.form.get('total_cozinha', '').replace(',', '.')
    obs       = request.form.get('obs', '').strip() or None

    item = db.execute('''
        SELECT ri.*, r.status FROM requisicao_itens ri
        JOIN requisicoes r ON r.id = ri.requisicao_id
        WHERE ri.id = ? AND ri.requisicao_id = ?
    ''', (item_id, id)).fetchone()
    if not item or item['status'] != 'aberto':
        db.close(); return ('', 204)

    total_consumido = db.execute(
        'SELECT COALESCE(SUM(qtd_consumida),0) FROM requisicao_consumos WHERE item_id=?',
        (item_id,)).fetchone()[0]
    saldo        = max(0.0, item['qtd_enviada'] - total_consumido)
    qtd_sobra    = max(0.0, min(saldo, qtd_sobra))
    qtd_consumida = max(0.0, saldo - qtd_sobra)

    db.execute('''INSERT INTO requisicao_consumos
        (requisicao_id, item_id, insumo_id, data, qtd_consumida, qtd_sobra, obs)
        VALUES (?,?,?,?,?,?,?)''',
        (id, item_id, item['insumo_id'], data, qtd_consumida, qtd_sobra, obs))

    try:
        total_cozinha = float(total_cozinha_str)
        db.execute('UPDATE insumos SET estoque_atual = MAX(0, ?) WHERE id=?',
                   (total_cozinha, item['insumo_id']))
    except (ValueError, TypeError):
        if qtd_consumida > 0:
            db.execute('UPDATE insumos SET estoque_atual = MAX(0, COALESCE(estoque_atual,0) - ?) WHERE id=?',
                       (qtd_consumida, item['insumo_id']))

    db.commit(); db.close()
    return ('', 204)


@app.route('/requisicoes/consumo/<int:cid>/excluir', methods=['POST'])
def requisicao_consumo_excluir(cid):
    db = get_db()
    rc = db.execute('SELECT * FROM requisicao_consumos WHERE id=?', (cid,)).fetchone()
    if not rc:
        db.close(); return ('', 404)
    if rc['qtd_consumida'] and rc['qtd_consumida'] > 0:
        db.execute('UPDATE insumos SET estoque_atual = COALESCE(estoque_atual,0) + ? WHERE id=?',
                   (rc['qtd_consumida'], rc['insumo_id']))
    db.execute('DELETE FROM requisicao_consumos WHERE id=?', (cid,))
    db.commit(); db.close()
    return ('', 204)


@app.route('/insumos/<int:id>/ajustar_estoque', methods=['POST'])
def insumo_ajustar_estoque(id):
    coz = request.form.get('estoque_atual', '').replace(',', '.')
    cen = request.form.get('estoque_central', '').replace(',', '.')
    try:
        coz = float(coz); cen = float(cen)
    except ValueError:
        return jsonify({'ok': False, 'erro': 'Valor inválido'}), 400
    db  = get_db()
    antes = db.execute('SELECT estoque_atual, estoque_central FROM insumos WHERE id=?', (id,)).fetchone()
    db.execute('UPDATE insumos SET estoque_atual=?, estoque_central=? WHERE id=?', (coz, cen, id))
    db.execute('''INSERT INTO insumo_ajuste_log
        (insumo_id, data, estoque_cozinha_antes, estoque_central_antes,
         estoque_cozinha_depois, estoque_central_depois)
        VALUES (?, date("now","localtime"), ?, ?, ?, ?)''',
        (id,
         antes['estoque_atual']   if antes else None,
         antes['estoque_central'] if antes else None,
         coz, cen))
    db.commit(); db.close()
    return jsonify({'ok': True, 'estoque_atual': coz, 'estoque_central': cen})


@app.route('/insumo_ajuste_log/<int:lid>/desfazer', methods=['POST'])
def insumo_ajuste_desfazer(lid):
    db  = get_db()
    row = db.execute('SELECT * FROM insumo_ajuste_log WHERE id=?', (lid,)).fetchone()
    if not row:
        db.close(); return jsonify({'ok': False, 'erro': 'Registro não encontrado'}), 404
    db.execute('UPDATE insumos SET estoque_atual=?, estoque_central=? WHERE id=?',
               (row['estoque_cozinha_antes'], row['estoque_central_antes'], row['insumo_id']))
    db.execute('DELETE FROM insumo_ajuste_log WHERE id=?', (lid,))
    db.commit(); db.close()
    return jsonify({'ok': True})


@app.route('/alertas')
def alertas():
    db = get_db()
    insumos = db.execute('''
        SELECT id, nome, unidade_uso, unidade_compra, qtd_por_embalagem,
               estoque_atual, estoque_central, estoque_minimo, fornecedor
        FROM insumos
        ORDER BY nome COLLATE NOCASE
    ''').fetchall()
    db.close()
    return render_template('alertas.html', insumos=insumos)

def count_alertas():
    try:
        db = get_db()
        n = db.execute('''SELECT COUNT(*) FROM insumos
                          WHERE estoque_minimo > 0 AND COALESCE(estoque_atual,0) < estoque_minimo''').fetchone()[0]
        db.close()
        return n
    except:
        return 0

@app.context_processor
def inject_alertas():
    return dict(n_alertas=count_alertas())

# ── Previsão de Produção ─────────────────────────────────────
def _norm_reg_name(s):
    import unicodedata as _ud, re as _re
    s = _ud.normalize('NFD', s or '')
    s = ''.join(c for c in s if _ud.category(c) != 'Mn')
    return _re.sub(r'\s+', ' ', s).upper().strip()

CONSOLIDADO_PATH = r'C:\Users\Selectus\Desktop\Analise Claude\Vendas diarias\Consolidado\Consolidado por PDV.xlsx'

def _ler_plano_producao():
    """Lê 'Plano Prod. Sem Feriado' abrindo Excel invisível para recalcular fórmulas."""
    try:
        return _ler_plano_via_com_fresh()
    except Exception:
        pass
    return _ler_plano_via_openpyxl_fallback()


def _ler_plano_via_com_fresh():
    """Abre Excel invisível, recalcula e lê — sempre a partir do arquivo em disco."""
    import win32com.client, pythoncom
    pythoncom.CoInitialize()
    xl = win32com.client.Dispatch("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    wb = None
    try:
        wb = xl.Workbooks.Open(CONSOLIDADO_PATH, ReadOnly=True, UpdateLinks=False)
        xl.CalculateFull()
        ws = wb.Sheets("Plano Prod. Sem Feriado")

        def _v(x):
            try: return float(x) if x is not None else 0.0
            except (TypeError, ValueError): return 0.0

        produtos = {}
        base_info = None
        cat_atual = ''
        last_row = ws.UsedRange.Row + ws.UsedRange.Rows.Count - 1

        for r in range(1, last_row + 1):
            row0 = r - 1
            cell_a = ws.Cells(r, 1).Value
            cell_b = ws.Cells(r, 2).Value

            if row0 == 1:
                base_info = str(cell_a or '').split('|')[0].replace('Base:', '').strip()
            if row0 < 4:
                continue
            if not cell_b:
                continue
            nome_str = str(cell_b).strip()
            if nome_str.startswith('  TOTAL') or nome_str.startswith('  '):
                continue
            if cell_a:
                cat_atual = str(cell_a).strip()

            # P=col16, Q=col17, R=col18 (COM 1-indexed) = TOTAL Sex, Seg, Qua
            sex = int(_v(ws.Cells(r, 16).Value))
            seg = int(_v(ws.Cells(r, 17).Value))
            qua = int(_v(ws.Cells(r, 18).Value))

            key = _norm_reg_name(nome_str)
            produtos[key] = {'nome': nome_str, 'cat': cat_atual,
                             'seg': seg, 'qua': qua, 'sex': sex}

        return produtos, base_info
    finally:
        if wb:
            wb.Close(SaveChanges=False)
        try: xl.Quit()
        except Exception: pass


def _ler_plano_via_openpyxl_fallback():
    """Fallback openpyxl — cache pode estar desatualizado para células de fórmula."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(CONSOLIDADO_PATH, read_only=True, data_only=True)
        ws = wb['Plano Prod. Sem Feriado']
    except Exception:
        return {}, None

    def _v(x):
        try: return float(x) if x is not None else 0.0
        except (TypeError, ValueError): return 0.0

    produtos = {}
    base_info = None
    cat_atual = ''

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 1:
            base_info = str(row[0] or '').split('|')[0].replace('Base:', '').strip()
        if i < 4:
            continue
        nome = row[1]
        if not nome:
            continue
        nome_str = str(nome).strip()
        if nome_str.startswith('  TOTAL') or nome_str.startswith('  '):
            continue
        if row[0]:
            cat_atual = str(row[0]).strip()
        sex = int(_v(row[15]))
        seg = int(_v(row[16]))
        qua = int(_v(row[17]))
        key = _norm_reg_name(nome_str)
        produtos[key] = {'nome': nome_str, 'cat': cat_atual,
                         'seg': seg, 'qua': qua, 'sex': sex}
    return produtos, base_info

def _get_prev_items(db, reg_id):
    if not reg_id:
        return {}
    rows = db.execute(
        'SELECT produto, categoria, qtd_planejada FROM registro_itens WHERE registro_id=?',
        (reg_id,)).fetchall()
    result = {}
    for r in rows:
        key = _norm_reg_name(r['produto'])
        result[key] = {'cat': r['categoria'] or '', 'qtd': r['qtd_planejada'] or 0,
                       'nome_orig': (r['produto'] or '').upper().strip()}
    return result

@app.route('/previsao', methods=['GET','POST'])
def previsao():
    EXCLUIR = {'MIX DE FRUTAS UVA E MELAO 150G'}

    plano, base_info = _ler_plano_producao()

    produtos = []
    for key, v in sorted(plano.items(), key=lambda x: (x[1]['cat'], x[0])):
        if key in EXCLUIR:
            continue
        p = dict(v)
        p['key'] = key
        p['prev_total'] = v['seg'] + v['qua'] + v['sex']
        p['slug'] = str(abs(hash(key)))
        produtos.append(p)

    return render_template('previsao.html', produtos=produtos, base_info=base_info)


@app.route('/previsao/necessidade', methods=['POST'])
def previsao_necessidade():
    norm_name = _norm_reg_name
    ALIASES = {
        'CUSCUZ':                           16,
        'MIX DE FRUTAS UVA E MANGA':        12,
        'SALADA TERIYAKI COM PROTEINA':      4,
        'SANDUICHE DE PEITO DE PERU':       11,
        'GELATINA DE MORANGO':              17,
        'GELATINA DE UVA':                  18,
        'OVERNIGHT 180G':                   31,
    }

    db = get_db()
    prods_all = db.execute('SELECT * FROM produtos ORDER BY nome').fetchall()
    prod_map = {norm_name(p['nome']): p['id'] for p in prods_all}
    for alias_key, pid in ALIASES.items():
        prod_map[norm_name(alias_key)] = pid

    slugs = [k[3:] for k in request.form if k.startswith('qs_')]
    matched_dia = {}
    matched_sem = {}
    sem_ficha = []

    for slug in slugs:
        nome = request.form.get(f'nom_{slug}', '')
        try: qty_seg = float(request.form.get(f'qs_{slug}', 0) or 0)
        except Exception: qty_seg = 0
        try: qty_qua = float(request.form.get(f'qq_{slug}', 0) or 0)
        except Exception: qty_qua = 0
        try: qty_sex = float(request.form.get(f'qx_{slug}', 0) or 0)
        except Exception: qty_sex = 0

        pid = prod_map.get(norm_name(nome))
        if pid:
            if qty_seg > 0:
                matched_dia[pid] = matched_dia.get(pid, 0) + qty_seg
            if qty_qua + qty_sex > 0:
                matched_sem[pid] = matched_sem.get(pid, 0) + qty_qua + qty_sex
        elif qty_seg + qty_qua + qty_sex > 0:
            sem_ficha.append({'nome': nome, 'qty_dia': qty_seg, 'qty_sem': qty_qua + qty_sex})

    # Lista do próximo dia (vs estoque atual)
    form_dia = {f'qty_{p["id"]}': matched_dia.get(p['id'], 0) for p in prods_all}
    lista_dia, custo_dia = _calcular_lista(prods_all, form_dia)
    lista_dia = [i for i in lista_dia if i['qtd_necessaria'] > 0]

    # Projetar estoque após o dia 1 (após compras + consumo)
    all_estoques = {r['id']: float(r['estoque_atual'] or 0)
                    for r in db.execute('SELECT id, estoque_atual FROM insumos').fetchall()}
    db.close()

    estoque_pos_dia1 = dict(all_estoques)
    for item in lista_dia:
        iid = item['id']
        est = item['estoque_disponivel']
        uso = item['qtd_necessaria']
        compra_uso = item['qtd_compra'] * item['fator_conversao']
        # Após consumir para o dia 1 e receber compras: sobra = est - uso + compra_uso
        estoque_pos_dia1[iid] = max(0.0, est - uso + compra_uso)

    # Lista do restante da semana (vs estoque projetado após dia 1)
    form_sem = {f'qty_{p["id"]}': matched_sem.get(p['id'], 0) for p in prods_all}
    lista_sem, custo_sem = _calcular_lista(prods_all, form_sem, estoque_override=estoque_pos_dia1)
    lista_sem = [i for i in lista_sem if i['qtd_necessaria'] > 0]

    resultado = {
        'lista_dia': lista_dia,
        'custo_dia': custo_dia,
        'lista_sem': lista_sem,
        'custo_sem': custo_sem,
        'custo_total': custo_dia + custo_sem,
        'sem_ficha': sem_ficha,
        'falta_preco': any(i['preco_compra'] is None for i in lista_dia + lista_sem),
    }
    return render_template('previsao_necessidade.html', resultado=resultado)


@app.route('/previsao/lista', methods=['POST'])
def previsao_lista():
    norm_name = _norm_reg_name

    # Mapeamento manual: nome normalizado do registro → id do produto
    ALIASES = {
        'CUSCUZ':                           16,  # Cuscuz carne do sol 170 gr
        'MIX DE FRUTAS UVA E MANGA':        12,  # Mix frutas uva + manga
        'SALADA TERIYAKI COM PROTEINA':      4,  # Salada teriyaki
        'SANDUICHE DE PEITO DE PERU':       11,  # Sanduíche misto com peito de peru
        'GELATINA DE MORANGO':              17,  # Sobremesa gelatina morango
        'GELATINA DE UVA':                  18,  # Sobremesa gelatina uva
        'OVERNIGHT 180G':                   31,  # Overnight Frutas vermelhas
    }

    db = get_db()
    prods_all = db.execute('SELECT * FROM produtos ORDER BY nome').fetchall()
    prod_map = {norm_name(p['nome']): p['id'] for p in prods_all}
    # Injeta aliases no mapa
    for alias_key, pid in ALIASES.items():
        prod_map[norm_name(alias_key)] = pid

    # Coleta quantidades submetidas: nom_<slug> + qty_<slug>
    slugs = [k[4:] for k in request.form if k.startswith('nom_')]
    matched = {}   # pid -> qty
    sem_ficha = [] # sem correspondência

    for slug in slugs:
        nome = request.form.get(f'nom_{slug}', '')
        try:
            qty = float(request.form.get(f'qty_{slug}', 0) or 0)
        except Exception:
            qty = 0
        if qty <= 0:
            continue
        pid = prod_map.get(norm_name(nome))
        if pid:
            matched[pid] = matched.get(pid, 0) + qty
        else:
            sem_ficha.append({'nome': nome, 'qty': qty})

    form_calc = {f'qty_{p["id"]}': matched.get(p['id'], 0) for p in prods_all}
    lista, custo_total = _calcular_lista(prods_all, form_calc)
    db.close()

    lista_filtrada = [i for i in lista if i['qtd_necessaria'] > 0]
    sobra_total = sum(i['sobra_valor'] for i in lista_filtrada if i.get('sobra_valor') is not None)
    resultado = {
        'lista': lista_filtrada,
        'custo_total': custo_total,
        'sobra_total': sobra_total,
        'falta_preco': any(i['preco_compra'] is None for i in lista_filtrada),
        'sem_ficha': sem_ficha,
    }
    return render_template('previsao_lista.html', resultado=resultado)

# ── Registro de Produção ─────────────────────────────────────
@app.route('/registros')
def registro_lista():
    db = get_db()
    regs = db.execute('''
        SELECT r.*, COUNT(i.id) as n_itens,
               SUM(CASE WHEN i.concluido=1 THEN 1 ELSE 0 END) as n_concluidos,
               COALESCE(SUM(i.qtd_planejada),0) as total_planejado
        FROM registros_producao r
        LEFT JOIN registro_itens i ON i.registro_id = r.id
        GROUP BY r.id ORDER BY r.data DESC, r.rodada
    ''').fetchall()
    db.close()
    return render_template('registro_lista.html', registros=regs)

@app.route('/registros/<int:id>', methods=['GET', 'POST'])
def registro_ver(id):
    db = get_db()
    reg = db.execute('SELECT * FROM registros_producao WHERE id=?', (id,)).fetchone()
    if not reg:
        db.close(); return redirect(url_for('registro_lista'))
    itens = db.execute('''SELECT ri.*,
           COALESCE((SELECT SUM(d.qtd_descartada) FROM descartes d WHERE d.registro_item_id=ri.id),0) as total_descarte,
           (SELECT GROUP_CONCAT(d.local||':'||d.qtd_descartada,' / ') FROM descartes d WHERE d.registro_item_id=ri.id) as desc_pdv
        FROM registro_itens ri WHERE ri.registro_id=?
        ORDER BY ri.ordem, ri.categoria, ri.produto''', (id,)).fetchall()
    db.close()
    cats = {}
    for it in itens:
        cat = it['categoria'] or 'Outros'
        cats.setdefault(cat, []).append(it)
    n_concluidos = sum(1 for i in itens if i['concluido'])

    unmatched = []
    db = get_db()
    prods = db.execute('SELECT * FROM produtos ORDER BY nome').fetchall()
    db.close()
    import unicodedata as _ud, re as _re2
    def norm(s):
        s = _ud.normalize('NFKD', s).encode('ascii','ignore').decode().lower().strip()
        return _re2.sub(r'\s+', ' ', s)
    prod_map = {norm(p['nome']): p for p in prods}
    form = {}
    for it in itens:
        qtd = it['qtd_planejada'] or 0
        if qtd <= 0:
            continue
        key = norm(it['produto'])
        found = prod_map.get(key)
        if not found:
            for pname, p in prod_map.items():
                if key in pname or pname in key:
                    found = p; break
        if not found:
            palavras = [w for w in key.split() if len(w) > 4]
            for pname, p in prod_map.items():
                if palavras and sum(1 for w in palavras if w in pname) >= max(1, len(palavras) - 1):
                    found = p; break
        if found:
            form[f'qty_{found["id"]}'] = qtd
        else:
            unmatched.append(it['produto'])
    lista, custo_total = _calcular_lista(prods, form)
    resultado = {'lista': lista, 'custo_total': custo_total,
                 'falta_preco': any(i['preco_compra'] is None for i in lista)}

    # Descartes por item e PDV
    db = get_db()
    desc_rows = db.execute(
        '''SELECT d.registro_item_id, d.local, SUM(d.qtd_descartada) as qtd
           FROM descartes d
           JOIN registro_itens ri ON ri.id = d.registro_item_id
           WHERE ri.registro_id=?
           GROUP BY d.registro_item_id, d.local
           ORDER BY d.registro_item_id, d.local''', (id,)
    ).fetchall()
    db.close()
    desc_map = {}  # {item_id: {pdv: qty}}
    for d in desc_rows:
        desc_map.setdefault(d['registro_item_id'], {})[d['local']] = d['qtd']

    # Vendas por produto e dia
    db = get_db()
    vendas_rows = db.execute(
        'SELECT * FROM registro_vendas WHERE registro_id=? ORDER BY data, produto', (id,)
    ).fetchall()
    db.close()
    # vendas_map: {produto -> [{'data', 'ptg', 'amaro', 'izabel', 'agnus', 'total'}]}
    vendas_map = {}
    vendas_datas = []
    for v in vendas_rows:
        prod = v['produto']
        d = v['data']
        if d not in vendas_datas:
            vendas_datas.append(d)
        vendas_map.setdefault(prod, []).append({
            'data': d, 'ptg': v['ptg'] or 0, 'amaro': v['amaro'] or 0,
            'izabel': v['izabel'] or 0, 'agnus': v['agnus'] or 0,
            'total': (v['ptg'] or 0) + (v['amaro'] or 0) + (v['izabel'] or 0) + (v['agnus'] or 0)
        })

    return render_template('registro_ver.html', reg=reg, itens=itens, cats=cats,
                           n_concluidos=n_concluidos, total_itens=len(itens),
                           resultado=resultado, unmatched=unmatched,
                           vendas_map=vendas_map, vendas_datas=vendas_datas,
                           desc_map=desc_map)

@app.route('/registros/<int:id>/exportar', methods=['GET'])
def registro_exportar(id):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash('openpyxl não instalado.', 'danger')
        return redirect(url_for('registro_ver', id=id))

    db = get_db()
    reg = db.execute('SELECT * FROM registros_producao WHERE id=?', (id,)).fetchone()
    itens = db.execute('SELECT * FROM registro_itens WHERE registro_id=? ORDER BY ordem, categoria, produto', (id,)).fetchall()
    prods = db.execute('SELECT * FROM produtos ORDER BY nome').fetchall()
    db.close()

    import unicodedata, re as _re
    def norm(s):
        s = unicodedata.normalize('NFKD', s).encode('ascii','ignore').decode().lower().strip()
        return _re.sub(r'\s+', ' ', s)
    prod_map = {norm(p['nome']): p for p in prods}
    form = {}
    for it in itens:
        qtd = it['qtd_planejada'] or 0
        if qtd <= 0: continue
        key = norm(it['produto'])
        found = prod_map.get(key)
        if not found:
            for pname, p in prod_map.items():
                if key in pname or pname in key:
                    found = p; break
        if found:
            form[f'qty_{found["id"]}'] = qtd
    lista, custo_total = _calcular_lista(prods, form)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Lista de Compras'
    verde_escuro = '375623'
    verde_claro  = 'E8F5E9'
    cinza_header = '1A3A5C'
    branco       = 'FFFFFF'
    def cell_fill(c): return PatternFill('solid', fgColor=c)
    def border_thin():
        s = Side(style='thin', color='CCCCCC')
        return Border(left=s, right=s, top=s, bottom=s)

    data_reg = reg['data'][8:10] + '/' + reg['data'][5:7] + '/' + reg['data'][0:4]
    ws.merge_cells('A1:I1')
    ws['A1'] = f'Lista de Compras — Registro {data_reg} (Rodada {reg["rodada"]})'
    ws['A1'].font = Font(bold=True, size=14, color=branco)
    ws['A1'].fill = cell_fill(cinza_header)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28
    ws.merge_cells('A2:I2')
    ws['A2'] = f'Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font = Font(italic=True, size=10, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 16

    headers = ['#', 'Insumo', 'Qtd. Necessária', 'Unid. Uso', 'Em Estoque',
               'Qtd. a Comprar', 'Unid. Compra', 'Preço / Unid. (R$)', 'Valor Previsto (R$)']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = Font(bold=True, color=branco, size=10)
        c.fill = cell_fill(verde_escuro)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border_thin()
    ws.row_dimensions[4].height = 32

    for i, item in enumerate(lista, 1):
        row = i + 4
        fill = cell_fill(verde_claro) if i % 2 == 0 else cell_fill(branco)
        if item['qtd_compra'] == 0: fill = cell_fill('F1F8E9')
        vals = [i, item['nome'], round(item['qtd_necessaria'], 2), item['unidade_uso'],
                round(item['estoque_disponivel'], 2) if item['estoque_disponivel'] else '—',
                item['qtd_compra'] if item['qtd_compra'] > 0 else 'Coberto',
                item['unidade_compra'],
                item['preco_compra'] if item['preco_compra'] else '—',
                round(item['valor'], 2) if item['preco_compra'] and item['qtd_compra'] > 0 else '—']
        aligns = ['center','left','right','center','right','right','center','right','right']
        for col, (val, aln) in enumerate(zip(vals, aligns), 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = fill
            c.alignment = Alignment(horizontal=aln, vertical='center')
            c.border = border_thin()
            c.font = Font(size=10)
            if col in (3, 5, 6) and isinstance(val, (int, float)): c.number_format = '#,##0.00'
            if col in (8, 9) and isinstance(val, float): c.number_format = 'R$ #,##0.0000'

    total_row = len(lista) + 5
    ws.merge_cells(f'A{total_row}:H{total_row}')
    ws[f'A{total_row}'] = 'VALOR TOTAL PREVISTO DE COMPRAS'
    ws[f'A{total_row}'].font = Font(bold=True, color=branco, size=11)
    ws[f'A{total_row}'].fill = cell_fill(verde_escuro)
    ws[f'A{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'I{total_row}'] = round(custo_total, 2)
    ws[f'I{total_row}'].font = Font(bold=True, color=branco, size=11)
    ws[f'I{total_row}'].fill = cell_fill(verde_escuro)
    ws[f'I{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'I{total_row}'].number_format = 'R$ #,##0.00'
    ws.row_dimensions[total_row].height = 24
    widths = [5, 40, 16, 12, 14, 16, 14, 20, 20]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'lista_compras_{data_reg.replace("/","_")}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/registros/<int:id>/item/<int:item_id>/salvar', methods=['POST'])
def registro_item_salvar(id, item_id):
    d = request.json or {}
    qtd = d.get('qtd_produzida')
    concluido = int(d.get('concluido', 0))
    obs = d.get('observacao', '')
    lote = d.get('lote', '')
    qtd_val = float(qtd) if str(qtd).strip() not in ('', 'null', 'None') else None
    def _int(v): return int(v) if str(v).strip() not in ('', 'null', 'None') else None
    port  = _int(d.get('qtd_portugues'))
    amaro = _int(d.get('qtd_amaro'))
    izabel= _int(d.get('qtd_izabel'))
    agnus = _int(d.get('qtd_agnus'))
    cris  = _int(d.get('qtd_cris'))
    env_port  = _int(d.get('env_portugues'))
    env_amaro = _int(d.get('env_amaro'))
    env_izabel= _int(d.get('env_izabel'))
    env_agnus = _int(d.get('env_agnus'))
    env_cris  = _int(d.get('env_cris'))
    db = get_db()
    db.execute('''UPDATE registro_itens
                  SET qtd_produzida=?, concluido=?, observacao=?, lote=?,
                      qtd_portugues=?, qtd_amaro=?, qtd_izabel=?, qtd_agnus=?, qtd_cris=?,
                      env_portugues=?, env_amaro=?, env_izabel=?, env_agnus=?, env_cris=?
                  WHERE id=? AND registro_id=?''',
               (qtd_val, concluido, obs, lote or None,
                port, amaro, izabel, agnus, cris,
                env_port, env_amaro, env_izabel, env_agnus, env_cris, item_id, id))
    db.commit()
    # Auto-atualizar status do registro se todos concluídos
    total = db.execute('SELECT COUNT(*) FROM registro_itens WHERE registro_id=?', (id,)).fetchone()[0]
    feitos = db.execute('SELECT COUNT(*) FROM registro_itens WHERE registro_id=? AND concluido=1', (id,)).fetchone()[0]
    if feitos > 0 and feitos < total:
        db.execute("UPDATE registros_producao SET status='em_producao' WHERE id=? AND status='planejado'", (id,))
    elif feitos == total and total > 0:
        db.execute("UPDATE registros_producao SET status='concluido' WHERE id=?", (id,))
    db.commit(); db.close()
    return jsonify(ok=True, n_concluidos=feitos, total=total)


@app.route('/registros/<int:id>/lote', methods=['POST'])
def registro_lote(id):
    lote = (request.json or {}).get('lote', '')
    db = get_db()
    db.execute('UPDATE registros_producao SET lote=? WHERE id=?', (lote or None, id))
    db.commit()
    db.close()
    return jsonify(ok=True)


@app.route('/registros/<int:id>/status', methods=['POST'])
def registro_status(id):
    status = request.form.get('status', '')
    obs = request.form.get('observacoes', '')
    db = get_db()
    db.execute('UPDATE registros_producao SET status=?, observacoes=? WHERE id=?', (status, obs, id))
    db.commit(); db.close()
    return redirect(url_for('registro_ver', id=id))

# ── CAFÉ ──────────────────────────────────────────────────────────────
NESTLE_BEBIDAS = [
    ('expresso',        'Expresso'),
    ('curto',           'Curto'),
    ('duplo',           'Duplo'),
    ('americano',       'Americano'),
    ('capuccino',       'Capuccino'),
    ('cafe_leite',      'Café c/ Leite'),
    ('cafe_leite_curto','Café c/L Curto'),
    ('achoc_kitkat',    'Achoc. KitKat'),
    ('capp_alpino',     'Capp. Alpino'),
    ('achoc_alpino',    'Achoc. Alpino'),
    ('mocaccino_2f',    'Mocaccino 2F'),
    ('achoc_2f',        'Achoc. 2F'),
]
BBC_BEBIDAS = [
    ('bbc_cafe_longo',   'Café Longo'),
    ('bbc_cafe_curto',   'Café Curto'),
    ('bbc_cafe_pingado', 'Café Pingado'),
    ('bbc_capuccino',    'Capuccino'),
    ('bbc_mocaccino',    'Mocaccino'),
    ('bbc_chocolate',    'Chocolate'),
    ('bbc_choc_leite',   'Choc. c/ Leite'),
    ('bbc_leite',        'Leite'),
    ('bbc_chap_canela',  'Chap. c/ Canela'),
]

def _calc_delta(rows, col):
    """Calcula delta (consumo diário) como diferença entre leituras consecutivas."""
    deltas = []
    prev = None
    for r in rows:
        v = r[col]
        if v is not None and prev is not None:
            d = v - prev
            deltas.append(d if d >= 0 else None)
        else:
            deltas.append(None)
        prev = v if v is not None else prev
    return deltas

CAFE_FICHA_INSUMOS = [
    {'nome': 'Café em Grãos',   'unid': 'PCT', 'tam_kg': 1.0,  'pct_cx': 1, 'rs_cx': 126.83},
    {'nome': 'Cappuccino (pó)', 'unid': 'PCT', 'tam_kg': 1.3,  'pct_cx': 1, 'rs_cx': 127.56},
    {'nome': 'Alpino',           'unid': 'PCT', 'tam_kg': 1.3,  'pct_cx': 1, 'rs_cx': 226.41},
    {'nome': 'Kit Kat',          'unid': 'PCT', 'tam_kg': 1.3,  'pct_cx': 1, 'rs_cx': 226.41},
    {'nome': 'Dois Frades',      'unid': 'PCT', 'tam_kg': 1.3,  'pct_cx': 1, 'rs_cx': 212.29},
    {'nome': 'Nescafé Alegria',  'unid': 'PCT', 'tam_kg': 1.3,  'pct_cx': 6, 'rs_cx': 640.32},
]

CAFE_FICHA_COMPOSICAO = [
    {'col': 'expresso',        'produto': 'Expresso',       'insumos': [{'nome': 'Café em Grãos',   'gram': 7.0,  'doses_pct': 142}]},
    {'col': 'curto',           'produto': 'Curto',          'insumos': [{'nome': 'Café em Grãos',   'gram': 7.0,  'doses_pct': 142}]},
    {'col': 'duplo',           'produto': 'Duplo',          'insumos': [{'nome': 'Café em Grãos',   'gram': 10.0, 'doses_pct': 100}]},
    {'col': 'americano',       'produto': 'Americano',      'insumos': [{'nome': 'Café em Grãos',   'gram': 7.0,  'doses_pct': 142}]},
    {'col': 'capuccino',       'produto': 'Cappuccino',     'insumos': [{'nome': 'Café em Grãos',   'gram': 4.5,  'doses_pct': 222},
                                                                         {'nome': 'Cappuccino (pó)', 'gram': 22.6, 'doses_pct': 57}]},
    {'col': 'cafe_leite',      'produto': 'Café com Leite', 'insumos': [{'nome': 'Cappuccino (pó)', 'gram': 23.7, 'doses_pct': 54}]},
    {'col': 'cafe_leite_curto','produto': 'Café c/L Curto', 'insumos': [{'nome': 'Cappuccino (pó)', 'gram': 11.0, 'doses_pct': 118}]},
    {'col': 'achoc_kitkat',    'produto': 'Achoc. KitKat',  'insumos': [{'nome': 'Kit Kat',          'gram': 25.0, 'doses_pct': 52}]},
    {'col': 'capp_alpino',     'produto': 'Capp. Alpino',   'insumos': [{'nome': 'Café em Grãos',   'gram': 4.5,  'doses_pct': 222},
                                                                         {'nome': 'Alpino',           'gram': 28.0, 'doses_pct': 46}]},
    {'col': 'achoc_alpino',    'produto': 'Achoc. Alpino',  'insumos': [{'nome': 'Alpino',           'gram': 28.0, 'doses_pct': 46}]},
    {'col': 'mocaccino_2f',    'produto': 'Mocaccino 2F',   'insumos': [{'nome': 'Café em Grãos',   'gram': 4.5,  'doses_pct': 222},
                                                                         {'nome': 'Dois Frades',      'gram': 25.0, 'doses_pct': 52}]},
    {'col': 'achoc_2f',        'produto': 'Achoc. 2F',      'insumos': [{'nome': 'Dois Frades',      'gram': 25.0, 'doses_pct': 52}]},
]

def _calc_custo_dose():
    """Calcula custo por dose a partir dos preços dos pacotes e composição."""
    preco_pct = {ins['nome']: ins['rs_cx'] / ins['pct_cx'] for ins in CAFE_FICHA_INSUMOS}
    custo = {}
    for item in CAFE_FICHA_COMPOSICAO:
        total = sum(preco_pct.get(ins['nome'], 0) / ins['doses_pct'] for ins in item['insumos'])
        custo[item['col']] = round(total, 4)
    return custo

CAFE_CUSTO_DOSE = _calc_custo_dose()
CAFE_PRECO_VENDA = {
    'expresso': 2.50, 'curto': 2.10, 'duplo': 2.90, 'americano': 3.10,
    'capuccino': 3.10, 'cafe_leite': 3.10, 'cafe_leite_curto': 3.10,
    'achoc_kitkat': 4.50, 'capp_alpino': 4.50, 'achoc_alpino': 4.10,
    'mocaccino_2f': 3.90, 'achoc_2f': 4.50,
}
CUSTO_COPO = 0.30

# Preços zerados por máquina (consumo interno / gratuito)
CAFE_PRECO_OVERRIDE = {
    3: {'expresso': 0, 'curto': 0},                      # Português CC
    4: {col: 0 for col, _ in NESTLE_BEBIDAS},            # Português Diretoria — tudo gratuito
}

CAFE_PDVS = {
    'Amaro':    {'label': 'Amaro',     'cor': '#1a3a5c', 'maquinas': [1, 2]},
    'Portugues':{'label': 'Português', 'cor': '#375623', 'maquinas': [3, 4, 5]},
    'Izabel':   {'label': 'Izabel',    'cor': '#5a1a3a', 'maquinas': [6]},
}

CAFE_LOCAIS_PDV    = ['Amaro', 'Portugues', 'Izabel']
CAFE_LOCAIS_TODOS  = ['Central', 'Amaro', 'Portugues', 'Izabel']
CAFE_ALERTA_PCT    = 0.5   # alerta de reposição abaixo de 0.5 pacote estimado
CAFE_ATENCAO_PCT   = 1.0   # atenção abaixo de 1 pacote


def _consumo_insumos_periodo(db, pdv, desde, ate=None):
    """Consumo de insumos (kg) de um PDV no período (desde exclusive, ate inclusive).
    Retorna {nome_insumo: kg}."""
    if ate is None:
        ate = date.today().isoformat()
    comp_map = {item['col']: item['insumos'] for item in CAFE_FICHA_COMPOSICAO}
    insumo_g = {}
    maquinas = db.execute(
        "SELECT id FROM cafe_maquinas WHERE pdv=? AND tipo!='bbc'", (pdv,)
    ).fetchall()
    for m in maquinas:
        mid = m['id']
        baseline = db.execute(
            "SELECT * FROM cafe_leituras WHERE maquina_id=? AND data<=? ORDER BY data DESC LIMIT 1",
            (mid, desde)
        ).fetchone()
        readings = db.execute(
            "SELECT * FROM cafe_leituras WHERE maquina_id=? AND data>? AND data<=? ORDER BY data ASC",
            (mid, desde, ate)
        ).fetchall()
        if not readings:
            continue
        prev = {col: ((baseline[col] or 0) if baseline else 0) for col, _ in NESTLE_BEBIDAS}
        for r in readings:
            for col, _ in NESTLE_BEBIDAS:
                v = r[col] or 0
                delta = max(v - prev[col], 0)
                prev[col] = v if v > 0 else prev[col]
                if delta > 0 and col in comp_map:
                    for ins in comp_map[col]:
                        insumo_g[ins['nome']] = insumo_g.get(ins['nome'], 0) + delta * ins['gram']
    return {nome: round(g / 1000, 4) for nome, g in insumo_g.items()}

@app.route('/cafe')
def cafe_lista():
    db = get_db()
    maquinas = db.execute("SELECT * FROM cafe_maquinas WHERE tipo != 'bbc' ORDER BY id").fetchall()
    resumo = []
    for m in maquinas:
        ultima = db.execute(
            'SELECT * FROM cafe_leituras WHERE maquina_id=? ORDER BY data DESC LIMIT 1', (m['id'],)
        ).fetchone()
        n_dias = db.execute(
            'SELECT COUNT(*) FROM cafe_leituras WHERE maquina_id=?', (m['id'],)
        ).fetchone()[0]
        primeira = db.execute(
            'SELECT data FROM cafe_leituras WHERE maquina_id=? ORDER BY data ASC LIMIT 1', (m['id'],)
        ).fetchone()
        bebidas = BBC_BEBIDAS if m['tipo'] == 'bbc' else NESTLE_BEBIDAS
        total_acum = None
        if ultima:
            vals = [ultima[col] for col, _ in bebidas if ultima[col] is not None]
            if vals:
                total_acum = sum(vals)
        # Consumo do último dia (delta: última - penúltima leitura)
        penultima = db.execute(
            'SELECT * FROM cafe_leituras WHERE maquina_id=? ORDER BY data DESC LIMIT 1 OFFSET 1', (m['id'],)
        ).fetchone()
        consumo_ultimo_dia = None
        if ultima and penultima:
            delta = sum(
                max(ultima[col] - penultima[col], 0)
                for col, _ in bebidas
                if ultima[col] is not None and penultima[col] is not None
            )
            consumo_ultimo_dia = delta
        resumo.append({
            'maquina': m,
            'ultima': ultima,
            'n_dias': n_dias,
            'primeira_data': primeira['data'] if primeira else None,
            'total_acum': total_acum,
            'consumo_ultimo_dia': consumo_ultimo_dia,
        })
    # Histórico diário consolidado (deltas de todas as máquinas)
    def _deltas_maq(rows_maq):
        prev = {col: None for col, _ in NESTLE_BEBIDAS}
        dias = {}
        for r in sorted(rows_maq, key=lambda x: x['data']):
            d_row = {}
            for col, _ in NESTLE_BEBIDAS:
                v, p = r[col], prev[col]
                d_row[col] = max(v - p, 0) if v is not None and p is not None else 0
                if v is not None: prev[col] = v
            dias[r['data']] = d_row
        return dias

    todas_leituras = db.execute(
        "SELECT * FROM cafe_leituras l JOIN cafe_maquinas m ON l.maquina_id=m.id WHERE m.tipo!='bbc' ORDER BY l.maquina_id, l.data"
    ).fetchall()
    por_maquina = {}
    for r in todas_leituras:
        por_maquina.setdefault(r['maquina_id'], []).append(r)

    maq_pdv = {r['maquina_id']: r['pdv'] for r in todas_leituras}

    hist_consolidado = {}  # {data: {col: total}}
    hist_por_pdv = {}      # {pdv: {data: {col: total}}}
    for mid, rows_m in por_maquina.items():
        pdv = maq_pdv.get(mid, 'Outro')
        for data, d_row in _deltas_maq(rows_m).items():
            if data not in hist_consolidado:
                hist_consolidado[data] = {col: 0 for col, _ in NESTLE_BEBIDAS}
            for col, _ in NESTLE_BEBIDAS:
                hist_consolidado[data][col] += d_row.get(col, 0)
            if pdv not in hist_por_pdv:
                hist_por_pdv[pdv] = {}
            if data not in hist_por_pdv[pdv]:
                hist_por_pdv[pdv][data] = {col: 0 for col, _ in NESTLE_BEBIDAS}
            for col, _ in NESTLE_BEBIDAS:
                hist_por_pdv[pdv][data][col] += d_row.get(col, 0)

    historico = [
        {'data': d, **hist_consolidado[d],
         'total': sum(hist_consolidado[d][col] for col, _ in NESTLE_BEBIDAS)}
        for d in sorted(hist_consolidado.keys(), reverse=True)
        if sum(hist_consolidado[d][col] for col, _ in NESTLE_BEBIDAS) > 0
    ]

    # Pacotes consumidos no último dia, por PDV
    insumo_map_fi = {i['nome']: i for i in CAFE_FICHA_INSUMOS}
    ultima_data_hist = historico[0]['data'] if historico else None
    pacotes_por_pdv = {}
    if ultima_data_hist:
        for pdv, hist_pdv in sorted(hist_por_pdv.items()):
            doses_pdv = hist_pdv.get(ultima_data_hist, {})
            insumo_g = {}
            for item in CAFE_FICHA_COMPOSICAO:
                doses = doses_pdv.get(item['col'], 0)
                for ins in item['insumos']:
                    insumo_g[ins['nome']] = insumo_g.get(ins['nome'], 0) + doses * ins['gram']
            pcts = []
            for nome, g in sorted(insumo_g.items()):
                if g == 0:
                    continue
                info = insumo_map_fi.get(nome, {})
                tam_g = info.get('tam_kg', 1.0) * 1000
                pcts.append({'nome': nome, 'g': round(g), 'pct': g / tam_g})
            if pcts:
                pacotes_por_pdv[pdv] = pcts
    pacotes_dia = []  # mantido para compatibilidade

    # Custo do dia (R$) — soma de pct * R$/PCT por insumo
    rs_pct_map = {i['nome']: i['rs_cx'] / i['pct_cx'] for i in CAFE_FICHA_INSUMOS}
    custo_dia_por_insumo = {}
    custo_dia_total = 0.0
    if ultima_data_hist:
        for pdv, pcts in pacotes_por_pdv.items():
            for p in pcts:
                rs = rs_pct_map.get(p['nome'], 0.0)
                custo = p['pct'] * rs
                custo_dia_por_insumo[p['nome']] = custo_dia_por_insumo.get(p['nome'], 0.0) + custo
                custo_dia_total += custo

    # Histórico diário de pacotes — todos os PDVs somados, por data
    all_dates_hist = set()
    for hist_pdv in hist_por_pdv.values():
        all_dates_hist.update(hist_pdv.keys())
    # todos os insumos que aparecem
    all_insumos_hist = sorted({ins['nome'] for item in CAFE_FICHA_COMPOSICAO for ins in item['insumos']})
    pacotes_historico = []
    for data in sorted(all_dates_hist, reverse=True):
        insumo_g = {}
        for pdv, hist_pdv in hist_por_pdv.items():
            doses_pdv = hist_pdv.get(data, {})
            for item in CAFE_FICHA_COMPOSICAO:
                doses = doses_pdv.get(item['col'], 0)
                for ins in item['insumos']:
                    insumo_g[ins['nome']] = insumo_g.get(ins['nome'], 0) + doses * ins['gram']
        itens = {}
        for nome, g in insumo_g.items():
            if g == 0:
                continue
            info = insumo_map_fi.get(nome, {})
            tam_g = (info.get('tam_kg', 1.0) or 1.0) * 1000
            itens[nome] = {'g': round(g), 'pct': g / tam_g}
        if itens:
            pacotes_historico.append({'data': data, 'itens': itens})

    # Tendência semanal por insumo
    from datetime import datetime, timedelta
    def _iso_week(data_str):
        d = datetime.strptime(data_str, '%Y-%m-%d').date()
        iso = d.isocalendar()
        return f'{iso[0]}-S{iso[1]:02d}'
    def _week_label(data_str):
        d = datetime.strptime(data_str, '%Y-%m-%d').date()
        # segunda-feira da semana
        seg = d - timedelta(days=d.weekday())
        dom = seg + timedelta(days=6)
        return f'{seg.strftime("%d/%m")}–{dom.strftime("%d/%m")}'

    semanas_pct = {}  # {iso_week: {insumo: pct_total}}
    semanas_label = {}
    for linha in pacotes_historico:
        wk = _iso_week(linha['data'])
        semanas_label[wk] = _week_label(linha['data'])
        if wk not in semanas_pct:
            semanas_pct[wk] = {}
        for nome, vals in linha['itens'].items():
            semanas_pct[wk][nome] = semanas_pct[wk].get(nome, 0.0) + vals['pct']

    semanas_ord = sorted(semanas_pct.keys())
    tendencia_semanas = []  # lista de {label, insumos: {nome: {pct, delta, trend}}}
    for i, wk in enumerate(semanas_ord):
        prev_wk = semanas_ord[i-1] if i > 0 else None
        insumos_t = {}
        for nome in all_insumos_hist:
            pct = semanas_pct[wk].get(nome, 0.0)
            prev_pct = semanas_pct[prev_wk].get(nome, 0.0) if prev_wk else None
            if prev_pct is not None and prev_pct > 0:
                delta_pct = ((pct - prev_pct) / prev_pct) * 100
                trend = 'up' if delta_pct > 5 else ('down' if delta_pct < -5 else 'stable')
            else:
                delta_pct = None
                trend = 'new'
            insumos_t[nome] = {'pct': pct, 'delta_pct': delta_pct, 'trend': trend}
        tendencia_semanas.append({'wk': wk, 'label': semanas_label[wk], 'insumos': insumos_t})
    tendencia_semanas = list(reversed(tendencia_semanas))  # mais recente primeiro

    # Consolidado geral
    ultima_data = max((r['ultima']['data'] for r in resumo if r['ultima']), default=None)
    # doses_hoje = total do hist_consolidado na ultima_data (mesma lógica da tabela)
    doses_hoje = historico[0]['total'] if historico else 0
    # total_periodo = soma de todos os deltas de todos os dias (historico já calculado)
    total_periodo = sum(h['total'] for h in historico)
    pdv_resumo  = {}
    for r in resumo:
        if not r['ultima'] or r['ultima']['data'] != ultima_data:
            continue
        pdv = r['maquina']['pdv']
        pdv_resumo.setdefault(pdv, {'hoje': 0})
        pdv_resumo[pdv]['hoje'] += r['consumo_ultimo_dia'] or 0

    db.close()
    return render_template('cafe_lista.html', resumo=resumo,
                           doses_hoje=doses_hoje, total_periodo=total_periodo,
                           ultima_data=ultima_data, pdv_resumo=pdv_resumo,
                           historico=historico, bebidas=NESTLE_BEBIDAS,
                           pacotes_dia=pacotes_dia, pacotes_por_pdv=pacotes_por_pdv,
                           pacotes_historico=pacotes_historico, all_insumos_hist=all_insumos_hist,
                           custo_dia_total=custo_dia_total, custo_dia_por_insumo=custo_dia_por_insumo,
                           tendencia_semanas=tendencia_semanas)

@app.route('/cafe/maquina/<int:mid>')
def cafe_maquina(mid):
    db = get_db()
    maquina = db.execute('SELECT * FROM cafe_maquinas WHERE id=?', (mid,)).fetchone()
    if not maquina:
        db.close(); return redirect(url_for('cafe_lista'))

    filtro_de  = request.args.get('de', '')
    filtro_ate = request.args.get('ate', '')

    # Buscar todos os registros para calcular deltas corretamente (inclui o anterior ao filtro)
    all_rows = db.execute(
        'SELECT * FROM cafe_leituras WHERE maquina_id=? ORDER BY data ASC', (mid,)
    ).fetchall()

    bebidas = BBC_BEBIDAS if maquina['tipo'] == 'bbc' else NESTLE_BEBIDAS
    dias_todos = []
    prev = {col: None for col, _ in bebidas}
    for r in all_rows:
        deltas = {}
        for col, label in bebidas:
            v = r[col]
            p = prev[col]
            if v is not None and p is not None:
                d = v - p
                deltas[col] = d if d >= 0 else None
            else:
                deltas[col] = None
            if v is not None:
                prev[col] = v
        dias_todos.append({'row': r, 'deltas': deltas})

    # Aplicar filtro de período
    dias = [d for d in dias_todos
            if (not filtro_de  or d['row']['data'] >= filtro_de)
            and (not filtro_ate or d['row']['data'] <= filtro_ate)]

    data_min = all_rows[0]['data'] if all_rows else ''
    data_max = all_rows[-1]['data'] if all_rows else ''

    db.close()
    return render_template('cafe_maquina.html', maquina=maquina, dias=dias,
                           bebidas=bebidas,
                           filtro_de=filtro_de, filtro_ate=filtro_ate,
                           data_min=data_min, data_max=data_max)

@app.route('/cafe/ficha')
def cafe_ficha():
    preco_pct = {ins['nome']: ins['rs_cx'] / ins['pct_cx'] for ins in CAFE_FICHA_INSUMOS}
    insumos = [{**ins, 'rs_pct': ins['rs_cx'] / ins['pct_cx']} for ins in CAFE_FICHA_INSUMOS]
    # Enriquecer composição com custo por insumo e total do produto
    composicao = []
    for item in CAFE_FICHA_COMPOSICAO:
        ins_enrich = []
        total = 0.0
        for ins in item['insumos']:
            custo = preco_pct.get(ins['nome'], 0) / ins['doses_pct']
            total += custo
            ins_enrich.append({**ins, 'custo_dose': custo})
        composicao.append({**item, 'insumos': ins_enrich, 'custo_total': total})
    bebidas_custo = [(col, label, CAFE_CUSTO_DOSE.get(col), CAFE_PRECO_VENDA.get(col))
                     for col, label in NESTLE_BEBIDAS]
    return render_template('cafe_ficha.html',
                           insumos=insumos,
                           composicao=composicao,
                           bebidas_custo=bebidas_custo,
                           custo_copo=CUSTO_COPO)

@app.route('/cafe/consolidado')
def cafe_consolidado():
    db = get_db()
    filtro_de  = request.args.get('de', '')
    filtro_ate = request.args.get('ate', '')

    resultado = {}
    for pdv_key, pdv_info in CAFE_PDVS.items():
        maquina_ids = pdv_info['maquinas']
        ph = ','.join('?' * len(maquina_ids))

        maquinas_info = {r['id']: dict(r) for r in db.execute(
            f'SELECT * FROM cafe_maquinas WHERE id IN ({ph})', maquina_ids).fetchall()}

        all_rows = db.execute(
            f'SELECT * FROM cafe_leituras WHERE maquina_id IN ({ph}) ORDER BY maquina_id, data ASC',
            maquina_ids).fetchall()

        # Calcular deltas por máquina
        deltas_por_maquina = {}
        for r in all_rows:
            mid = r['maquina_id']
            if mid not in deltas_por_maquina:
                deltas_por_maquina[mid] = {'prev': {col: None for col, _ in NESTLE_BEBIDAS}, 'dias': {}}
            state = deltas_por_maquina[mid]
            d_row = {}
            for col, _ in NESTLE_BEBIDAS:
                v = r[col]
                p = state['prev'][col]
                if v is not None and p is not None:
                    d = v - p
                    d_row[col] = d if d >= 0 else 0
                else:
                    d_row[col] = 0
                if v is not None:
                    state['prev'][col] = v
            state['dias'][r['data']] = d_row

        todas_datas = sorted({r['data'] for r in all_rows})

        def _filtrar_dias(dias_dict):
            result = []
            for data in todas_datas:
                if filtro_de and data < filtro_de: continue
                if filtro_ate and data > filtro_ate: continue
                if data not in dias_dict: continue
                totais = dias_dict[data]
                result.append({'data': data, 'totais': totais, 'total': sum(totais.values())})
            return result

        # Dados por máquina individual
        maquinas_dados = []
        for mid in maquina_ids:
            if mid not in deltas_por_maquina:
                maquinas_dados.append({'maquina': maquinas_info.get(mid, {}), 'dias': [], 'acum': {col: 0 for col, _ in NESTLE_BEBIDAS}})
                continue
            state = deltas_por_maquina[mid]
            acum_m = {col: (state['prev'][col] or 0) for col, _ in NESTLE_BEBIDAS}
            maquinas_dados.append({'maquina': maquinas_info[mid], 'dias': _filtrar_dias(state['dias']), 'acum': acum_m})

        # Consolidado (soma de todas as máquinas)
        consolidado_dias_raw = {}
        for mid, state in deltas_por_maquina.items():
            for data, d_row in state['dias'].items():
                if data not in consolidado_dias_raw:
                    consolidado_dias_raw[data] = {col: 0 for col, _ in NESTLE_BEBIDAS}
                for col, _ in NESTLE_BEBIDAS:
                    consolidado_dias_raw[data][col] += d_row.get(col, 0)

        acum = {col: 0 for col, _ in NESTLE_BEBIDAS}
        for mid, state in deltas_por_maquina.items():
            for col, _ in NESTLE_BEBIDAS:
                if state['prev'][col] is not None:
                    acum[col] += state['prev'][col]

        resultado[pdv_key] = {
            'info': pdv_info,
            'maquinas': maquinas_dados,
            'dias': _filtrar_dias(consolidado_dias_raw),
            'acum': acum,
            'data_min': todas_datas[0] if todas_datas else '',
            'data_max': todas_datas[-1] if todas_datas else '',
        }

    todas = [v['data_min'] for v in resultado.values() if v['data_min']]
    todas_max = [v['data_max'] for v in resultado.values() if v['data_max']]
    data_min = min(todas) if todas else ''
    data_max = max(todas_max) if todas_max else ''

    db.close()
    return render_template('cafe_consolidado.html',
                           resultado=resultado,
                           bebidas=NESTLE_BEBIDAS,
                           filtro_de=filtro_de, filtro_ate=filtro_ate,
                           data_min=data_min, data_max=data_max)

@app.route('/cafe/financeiro')
def cafe_financeiro():
    db = get_db()
    filtro_de  = request.args.get('de', '')
    filtro_ate = request.args.get('ate', '')

    def _deltas_maquina(rows_maq):
        """Retorna {data: {col: delta}} para uma lista de rows de uma máquina."""
        prev = {col: None for col, _ in NESTLE_BEBIDAS}
        dias = {}
        for r in sorted(rows_maq, key=lambda x: x['data']):
            d_row = {}
            for col, _ in NESTLE_BEBIDAS:
                v = r[col]
                p = prev[col]
                if v is not None and p is not None:
                    d = v - p
                    d_row[col] = d if d >= 0 else 0
                else:
                    d_row[col] = 0
                if v is not None:
                    prev[col] = v
            dias[r['data']] = d_row
        return dias

    def _stats(dias_dict, filtro_de, filtro_ate):
        """Calcula total doses por bebida e n_dias com dados no período."""
        total = {col: 0 for col, _ in NESTLE_BEBIDAS}
        n_dias = 0
        for data, d_row in dias_dict.items():
            if filtro_de and data < filtro_de: continue
            if filtro_ate and data > filtro_ate: continue
            if sum(d_row.values()) == 0: continue
            n_dias += 1
            for col, _ in NESTLE_BEBIDAS:
                total[col] += d_row.get(col, 0)
        return total, n_dias

    resultado = {}
    for pdv_key, pdv_info in CAFE_PDVS.items():
        maquina_ids = pdv_info['maquinas']
        ph = ','.join('?' * len(maquina_ids))

        maquinas_info = {r['id']: dict(r) for r in db.execute(
            f'SELECT * FROM cafe_maquinas WHERE id IN ({ph})', maquina_ids).fetchall()}

        all_rows = db.execute(
            f'SELECT * FROM cafe_leituras WHERE maquina_id IN ({ph}) ORDER BY maquina_id, data ASC',
            maquina_ids).fetchall()

        rows_por_maquina = {}
        for r in all_rows:
            rows_por_maquina.setdefault(r['maquina_id'], []).append(r)

        def _financeiro(total_doses, n_dias, mid=None):
            override = CAFE_PRECO_OVERRIDE.get(mid, {})
            rows = []
            totais = {k: 0 for k in ['receita_periodo', 'custo_insumo_periodo', 'custo_copo_periodo',
                                      'custo_total_periodo', 'margem_periodo',
                                      'receita_dia', 'custo_insumo_dia', 'custo_copo_dia',
                                      'custo_total_dia', 'margem_dia']}
            for col, label in NESTLE_BEBIDAS:
                doses_total = total_doses.get(col, 0)
                if doses_total == 0:
                    continue
                avg   = doses_total / n_dias if n_dias else 0
                preco = override.get(col, CAFE_PRECO_VENDA.get(col, 0))
                c_ins = CAFE_CUSTO_DOSE.get(col, 0)
                c_cop = CUSTO_COPO
                c_tot = c_ins + c_cop

                receita_periodo       = doses_total * preco
                custo_insumo_periodo  = doses_total * c_ins
                custo_copo_periodo    = doses_total * c_cop
                custo_total_periodo   = doses_total * c_tot
                margem_periodo        = receita_periodo - custo_total_periodo

                receita_dia           = avg * preco
                custo_insumo_dia      = avg * c_ins
                custo_copo_dia        = avg * c_cop
                custo_total_dia       = avg * c_tot
                margem_dia            = receita_dia - custo_total_dia

                margem_pct = (margem_periodo / receita_periodo * 100) if receita_periodo > 0 else 0
                rows.append({
                    'col': col, 'label': label,
                    'doses_total': doses_total, 'doses_dia': avg,
                    'receita_periodo': receita_periodo,
                    'custo_insumo_periodo': custo_insumo_periodo,
                    'custo_copo_periodo':   custo_copo_periodo,
                    'custo_total_periodo':  custo_total_periodo,
                    'margem_periodo':       margem_periodo,
                    'receita_dia':          receita_dia,
                    'custo_insumo_dia':     custo_insumo_dia,
                    'custo_copo_dia':       custo_copo_dia,
                    'custo_total_dia':      custo_total_dia,
                    'margem_dia':           margem_dia,
                    'margem_pct':           margem_pct,
                })
                totais['receita_periodo']      += receita_periodo
                totais['custo_insumo_periodo'] += custo_insumo_periodo
                totais['custo_copo_periodo']   += custo_copo_periodo
                totais['custo_total_periodo']  += custo_total_periodo
                totais['margem_periodo']       += margem_periodo
                totais['receita_dia']          += receita_dia
                totais['custo_insumo_dia']     += custo_insumo_dia
                totais['custo_copo_dia']       += custo_copo_dia
                totais['custo_total_dia']      += custo_total_dia
                totais['margem_dia']           += margem_dia
            t = totais
            t['margem_pct'] = (t['margem_periodo'] / t['receita_periodo'] * 100) if t['receita_periodo'] > 0 else 0
            return rows, totais

        maquinas_dados = []
        total_consolidado = {col: 0 for col, _ in NESTLE_BEBIDAS}
        n_dias_max = 0
        for mid in maquina_ids:
            rows_m = rows_por_maquina.get(mid, [])
            dias_m = _deltas_maquina(rows_m)
            total_m, n_m = _stats(dias_m, filtro_de, filtro_ate)
            fin_rows, fin_totais = _financeiro(total_m, n_m, mid=mid)
            maquinas_dados.append({
                'maquina': maquinas_info.get(mid, {}),
                'n_dias': n_m,
                'rows': fin_rows,
                'totais': fin_totais,
            })
            for col, _ in NESTLE_BEBIDAS:
                total_consolidado[col] += total_m.get(col, 0)
            if n_m > n_dias_max:
                n_dias_max = n_m

        fin_rows_c, fin_totais_c = _financeiro(total_consolidado, n_dias_max)

        # Datas para o filtro
        todas_datas = sorted({r['data'] for r in all_rows})
        resultado[pdv_key] = {
            'info': pdv_info,
            'maquinas': maquinas_dados,
            'consolidado': {'n_dias': n_dias_max, 'rows': fin_rows_c, 'totais': fin_totais_c},
            'data_min': todas_datas[0] if todas_datas else '',
            'data_max': todas_datas[-1] if todas_datas else '',
        }

    todas = [v['data_min'] for v in resultado.values() if v['data_min']]
    todas_max = [v['data_max'] for v in resultado.values() if v['data_max']]
    data_min = min(todas) if todas else ''
    data_max = max(todas_max) if todas_max else ''

    db.close()
    return render_template('cafe_financeiro.html',
                           resultado=resultado,
                           filtro_de=filtro_de, filtro_ate=filtro_ate,
                           data_min=data_min, data_max=data_max)


@app.route('/cafe/projecao')
def cafe_projecao():
    db = get_db()

    def _deltas_maquina(rows_maq):
        prev = {col: None for col, _ in NESTLE_BEBIDAS}
        dias = {}
        for r in sorted(rows_maq, key=lambda x: x['data']):
            d_row = {}
            for col, _ in NESTLE_BEBIDAS:
                v, p = r[col], prev[col]
                if v is not None and p is not None:
                    d = v - p
                    d_row[col] = d if d >= 0 else 0
                else:
                    d_row[col] = 0
                if v is not None:
                    prev[col] = v
            dias[r['data']] = d_row
        return dias

    def _stats(dias_dict):
        total = {col: 0 for col, _ in NESTLE_BEBIDAS}
        n = 0
        for d_row in dias_dict.values():
            if sum(d_row.values()) == 0:
                continue
            n += 1
            for col, _ in NESTLE_BEBIDAS:
                total[col] += d_row.get(col, 0)
        return total, n

    # Índices rápidos
    insumo_map  = {i['nome']: i for i in CAFE_FICHA_INSUMOS}
    comp_map    = {item['col']: item['insumos'] for item in CAFE_FICHA_COMPOSICAO}

    resultado   = {}
    consolidado = {i['nome']: 0.0 for i in CAFE_FICHA_INSUMOS}  # g/semana total

    for pdv_key, pdv_info in CAFE_PDVS.items():
        maquina_ids = pdv_info['maquinas']
        ph = ','.join('?' * len(maquina_ids))
        rows_all = db.execute(
            f'SELECT l.* FROM cafe_leituras l '
            f'JOIN cafe_maquinas m ON l.maquina_id = m.id '
            f'WHERE l.maquina_id IN ({ph}) AND m.tipo != "bbc"',
            maquina_ids,
        ).fetchall()

        rows_por_maquina = {}
        for r in rows_all:
            rows_por_maquina.setdefault(r['maquina_id'], []).append(r)

        total_doses = {col: 0 for col, _ in NESTLE_BEBIDAS}
        n_dias_max  = 0
        for mid in maquina_ids:
            dias_m = _deltas_maquina(rows_por_maquina.get(mid, []))
            tm, nm = _stats(dias_m)
            for col in total_doses:
                total_doses[col] += tm.get(col, 0)
            n_dias_max = max(n_dias_max, nm)

        if n_dias_max == 0:
            continue

        avg_dia = {col: total_doses[col] / n_dias_max for col in total_doses}

        # Insumo → g/dia acumulado para este PDV
        insumos_pdv = {}
        for col, _ in NESTLE_BEBIDAS:
            avg = avg_dia.get(col, 0)
            if avg == 0:
                continue
            for ins in comp_map.get(col, []):
                nome = ins['nome']
                g_dia = avg * ins['gram']
                insumos_pdv.setdefault(nome, 0.0)
                insumos_pdv[nome] += g_dia

        # Enriquece com g/semana e PCT/semana
        linhas = []
        for nome, g_dia in sorted(insumos_pdv.items()):
            info  = insumo_map.get(nome, {})
            pct_g = (info.get('tam_kg', 1.0)) * 1000
            g_sem = g_dia * 7
            rs_pct = info.get('rs_cx', 0) / max(info.get('pct_cx', 1), 1)
            pct_c = math.ceil(g_sem / pct_g)
            linhas.append({
                'nome':        nome,
                'tam_kg':      info.get('tam_kg', 1.0),
                'g_dia':       g_dia,
                'g_semana':    g_sem,
                'pct_raw':     g_sem / pct_g,
                'pct_ceil':    pct_c,
                'valor_total': pct_c * rs_pct,
            })
            consolidado[nome] = consolidado.get(nome, 0.0) + g_sem

        resultado[pdv_key] = {
            'info':    pdv_info,
            'n_dias':  n_dias_max,
            'linhas':  linhas,
            'avg_dia': avg_dia,
        }

    # Consolida todos os PDVs
    consolidado_linhas = []
    for nome, g_sem in sorted(consolidado.items()):
        if g_sem == 0:
            continue
        info  = insumo_map.get(nome, {})
        pct_g = info.get('tam_kg', 1.0) * 1000
        rs_pct = info.get('rs_cx', 0) / max(info.get('pct_cx', 1), 1)
        pct_c = math.ceil(g_sem / pct_g)
        consolidado_linhas.append({
            'nome':        nome,
            'tam_kg':      info.get('tam_kg', 1.0),
            'g_semana':    g_sem,
            'pct_raw':     g_sem / pct_g,
            'pct_ceil':    pct_c,
            'valor_total': pct_c * rs_pct,
        })

    total_valor = sum(l['valor_total'] for l in consolidado_linhas)

    db.close()
    return render_template('cafe_projecao.html',
                           resultado=resultado,
                           consolidado=consolidado_linhas,
                           total_valor=total_valor)

# ─── Consolidado por PDV ─────────────────────────────────────────────────────
CONSOLIDADO_PATH = r'C:\Users\Selectus\Desktop\Consolidado por PDV.xlsx'
_consolidado_sheets_cache = None
_consolidado_mtime_cache  = None
_consolidado_color_cache  = {}   # sheet_name → {mtime, rows, total_rows, total_cols}

# Abas que devem preservar as cores do Excel
COLORED_SHEETS = {'Alerta Sem Venda', 'Dashboard Alertas'}

def _get_sheet_names():
    global _consolidado_sheets_cache, _consolidado_mtime_cache
    try:
        mtime = os.path.getmtime(CONSOLIDADO_PATH)
        if _consolidado_sheets_cache is None or mtime != _consolidado_mtime_cache:
            import openpyxl
            wb = openpyxl.load_workbook(CONSOLIDADO_PATH, read_only=True)
            _consolidado_sheets_cache = wb.sheetnames
            _consolidado_mtime_cache  = mtime
            wb.close()
    except Exception:
        _consolidado_sheets_cache = []
    return _consolidado_sheets_cache or []

def _rgb_to_css(rgb):
    """Converte 'FFB71C1C' → '#B71C1C'. Retorna None se transparente."""
    if not rgb or len(rgb) < 8 or rgb[:2] == '00':
        return None
    return '#' + rgb[2:]

def _read_sheet_with_colors(sheet_name, max_rows=700):
    global _consolidado_color_cache
    try:
        mtime = os.path.getmtime(CONSOLIDADO_PATH)
    except Exception:
        mtime = 0

    cached = _consolidado_color_cache.get(sheet_name)
    if cached and cached['mtime'] == mtime:
        return cached['rows'], cached['total_rows'], cached['total_cols']

    import openpyxl
    wb = openpyxl.load_workbook(CONSOLIDADO_PATH, data_only=True)
    ws = wb[sheet_name]
    total_rows = ws.max_row or 0
    total_cols = ws.max_column or 0

    result_rows = []
    max_used_col = 0

    for r_idx, row in enumerate(ws.iter_rows(max_row=max_rows)):
        row_data = []
        has_content = False
        for c_idx, cell in enumerate(row):
            v = cell.value
            bg = fg = None
            bold = False
            try:
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.type == 'rgb':
                    bg = _rgb_to_css(cell.fill.fgColor.rgb)
            except Exception:
                pass
            try:
                if cell.font:
                    if cell.font.color and cell.font.color.type == 'rgb':
                        fg = _rgb_to_css(cell.font.color.rgb)
                    bold = bool(cell.font.bold)
            except Exception:
                pass

            cell_v = str(v).replace('\n', ' ') if v is not None else ''
            row_data.append({'v': cell_v, 'bg': bg, 'fg': fg, 'b': bold})
            if cell_v or bg:
                has_content = True
                max_used_col = max(max_used_col, c_idx + 1)

        if has_content:
            result_rows.append(row_data)

    # Corta colunas vazias no final
    if max_used_col:
        result_rows = [r[:max_used_col] for r in result_rows]

    wb.close()
    _consolidado_color_cache[sheet_name] = {
        'mtime': mtime, 'rows': result_rows,
        'total_rows': total_rows, 'total_cols': total_cols,
    }
    return result_rows, total_rows, total_cols

def _read_sheet(sheet_name, max_rows=500, max_cols=60):
    try:
        df = pd.read_excel(CONSOLIDADO_PATH, sheet_name=sheet_name,
                           header=0, nrows=max_rows + 1, engine='openpyxl')
        df = df.dropna(how='all', axis=0).dropna(how='all', axis=1).reset_index(drop=True)
        total_rows = len(df)
        total_cols = len(df.columns)
        truncated_cols = total_cols > max_cols
        if truncated_cols:
            df = df.iloc[:, :max_cols]
        cols = [str(c) for c in df.columns]
        rows = []
        for _, row in df.iterrows():
            cells = []
            for v in row:
                try:
                    if pd.isna(v):
                        cells.append('')
                        continue
                except Exception:
                    pass
                if isinstance(v, float) and v == int(v):
                    cells.append(str(int(v)))
                else:
                    cells.append(str(v))
            rows.append(cells)
        return cols, rows, total_rows, total_cols, truncated_cols
    except Exception as e:
        return [], [], 0, 0, False

@app.route('/consolidado')
def consolidado():
    sheet_names = _get_sheet_names()
    return render_template('consolidado.html', sheet_names=sheet_names)

@app.route('/consolidado/aba')
def consolidado_aba():
    sheet_names = _get_sheet_names()
    name = request.args.get('nome', '')
    if name not in sheet_names:
        return jsonify({'error': 'Aba não encontrada'}), 404

    if name in COLORED_SHEETS:
        rows, total_rows, total_cols = _read_sheet_with_colors(name)
        return jsonify({
            'nome': name, 'has_colors': True,
            'rows': rows,
            'total_rows': total_rows, 'total_cols': total_cols,
            'shown_rows': len(rows), 'cols': [], 'truncated_cols': False,
        })

    cols, rows, total_rows, total_cols, trunc_cols = _read_sheet(name)
    return jsonify({
        'nome': name, 'has_colors': False,
        'cols': cols, 'rows': rows,
        'total_rows': total_rows, 'total_cols': total_cols,
        'truncated_cols': trunc_cols, 'shown_rows': len(rows),
    })

# ─────────────── CAFÉ — ESTOQUE ──────────────────────────────────────────────

def _init_estoque_tables(db):
    db.execute('''CREATE TABLE IF NOT EXISTS cafe_estoque_entradas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        data TEXT NOT NULL,
        origem TEXT NOT NULL DEFAULT 'Fornecedor',
        local TEXT NOT NULL DEFAULT 'Central',
        insumo TEXT NOT NULL,
        qtd_kg REAL NOT NULL,
        obs TEXT,
        criado_em TEXT DEFAULT (datetime('now'))
    )''')
    db.execute('''CREATE TABLE IF NOT EXISTS cafe_estoque_contagens (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        data TEXT NOT NULL,
        local TEXT NOT NULL,
        insumo TEXT NOT NULL,
        qtd_kg REAL NOT NULL,
        obs TEXT,
        criado_em TEXT DEFAULT (datetime('now'))
    )''')
    db.commit()


def _calc_estoque_status(db):
    """Retorna status de estoque por insumo: estimado, variância, alerta."""
    hoje = date.today().isoformat()
    resultado = {}
    for ins in CAFE_FICHA_INSUMOS:
        nome = ins['nome']
        tam_kg = ins['tam_kg']

        # Última contagem como âncora (qualquer local, mais recente)
        ult = db.execute(
            "SELECT * FROM cafe_estoque_contagens WHERE insumo=? ORDER BY data DESC, id DESC LIMIT 1",
            (nome,)
        ).fetchone()
        desde = ult['data'] if ult else '2020-01-01'
        base_kg = float(ult['qtd_kg']) if ult else 0.0

        # Apenas compras de fornecedor adicionam ao estoque global (transferências são movimento interno)
        entradas_kg = db.execute(
            "SELECT COALESCE(SUM(qtd_kg),0) FROM cafe_estoque_entradas WHERE insumo=? AND data>=? AND (origem='Fornecedor' OR origem IS NULL)",
            (nome, desde)
        ).fetchone()[0] or 0.0

        # Consumo calculado desde a âncora (todos PDVs)
        consumo_kg = sum(
            _consumo_insumos_periodo(db, pdv, desde, hoje).get(nome, 0.0)
            for pdv in CAFE_LOCAIS_PDV
        )

        estimado_kg  = base_kg + entradas_kg - consumo_kg
        estimado_pct = estimado_kg / tam_kg if tam_kg else 0.0

        # Variância entre as duas contagens mais recentes
        variancia_pct = None
        pen = db.execute(
            "SELECT * FROM cafe_estoque_contagens WHERE insumo=? ORDER BY data DESC, id DESC LIMIT 1 OFFSET 1",
            (nome,)
        ).fetchone()
        if ult and pen:
            desde_p = pen['data']
            ate_p   = ult['data']
            base_p  = float(pen['qtd_kg'])
            ent_p   = db.execute(
                "SELECT COALESCE(SUM(qtd_kg),0) FROM cafe_estoque_entradas WHERE insumo=? AND data>=? AND data<=? AND (origem='Fornecedor' OR origem IS NULL)",
                (nome, desde_p, ate_p)
            ).fetchone()[0] or 0.0
            cons_p  = sum(
                _consumo_insumos_periodo(db, pdv, desde_p, ate_p).get(nome, 0.0)
                for pdv in CAFE_LOCAIS_PDV
            )
            estimado_em_ult = base_p + ent_p - cons_p
            real_ult = float(ult['qtd_kg'])
            if real_ult > 0:
                variancia_pct = (estimado_em_ult - real_ult) / real_ult * 100

        # Contagens por local (mais recente de cada)
        conts_local = {}
        for local in CAFE_LOCAIS_TODOS:
            c = db.execute(
                "SELECT * FROM cafe_estoque_contagens WHERE insumo=? AND local=? ORDER BY data DESC LIMIT 1",
                (nome, local)
            ).fetchone()
            if c:
                conts_local[local] = c

        resultado[nome] = {
            'tam_kg':         tam_kg,
            'base_kg':        round(base_kg, 3),
            'desde':          desde,
            'local_ancora':   ult['local'] if ult else None,
            'entradas_kg':    round(entradas_kg, 3),
            'consumo_kg':     round(consumo_kg, 3),
            'estimado_kg':    round(estimado_kg, 3),
            'estimado_pct':   round(estimado_pct, 2),
            'variancia_pct':  round(variancia_pct, 1) if variancia_pct is not None else None,
            'alerta':         estimado_pct < CAFE_ALERTA_PCT,
            'atencao':        CAFE_ALERTA_PCT <= estimado_pct < CAFE_ATENCAO_PCT,
            'conts_local':    conts_local,
        }
    return resultado


def _calc_estoque_por_local(db):
    """Estoque estimado por local × insumo.
    Fórmula: última contagem do local + entradas para o local - saídas do local - consumo (PDVs).
    Se o local não tem contagem própria, usa a data da contagem mais antiga do sistema com base 0."""
    hoje = date.today().isoformat()
    # Data mínima global: primeira contagem registrada no sistema (zeragem)
    row_min = db.execute("SELECT MIN(data) FROM cafe_estoque_contagens").fetchone()
    data_inicio = row_min[0] if row_min and row_min[0] else hoje
    resultado = {}
    for local in CAFE_LOCAIS_TODOS:
        resultado[local] = {}
        for ins in CAFE_FICHA_INSUMOS:
            nome   = ins['nome']
            tam_kg = ins['tam_kg']
            # Âncora: última contagem física DESTE local (ou base zero desde o início do sistema)
            ult = db.execute(
                "SELECT * FROM cafe_estoque_contagens WHERE insumo=? AND local=? ORDER BY data DESC, id DESC LIMIT 1",
                (nome, local)
            ).fetchone()
            desde   = ult['data']  if ult else data_inicio
            base_kg = float(ult['qtd_kg']) if ult else 0.0
            # Entradas para este local (compras se for Central; transferências recebidas)
            ent_in = db.execute(
                "SELECT COALESCE(SUM(qtd_kg),0) FROM cafe_estoque_entradas WHERE insumo=? AND local=? AND data>=?",
                (nome, local, desde)
            ).fetchone()[0] or 0.0
            # Saídas deste local (transferências enviadas)
            sai_out = db.execute(
                "SELECT COALESCE(SUM(qtd_kg),0) FROM cafe_estoque_entradas WHERE insumo=? AND origem=? AND data>=?",
                (nome, local, desde)
            ).fetchone()[0] or 0.0
            # Consumo calculado pelas máquinas (apenas PDVs)
            consumo = 0.0
            if local in CAFE_LOCAIS_PDV:
                consumo = _consumo_insumos_periodo(db, local, desde, hoje).get(nome, 0.0)
            estimado_kg  = base_kg + ent_in - sai_out - consumo
            estimado_pct = estimado_kg / tam_kg if tam_kg else 0.0
            resultado[local][nome] = {
                'estimado_kg':  round(estimado_kg, 3),
                'estimado_pct': round(estimado_pct, 2),
                'desde': desde,
                'tem_contagem': ult is not None,
                'alerta':  estimado_pct < CAFE_ALERTA_PCT,
                'atencao': CAFE_ALERTA_PCT <= estimado_pct < CAFE_ATENCAO_PCT,
            }
    return resultado


@app.route('/cafe/estoque')
def cafe_estoque():
    db = get_db()
    _init_estoque_tables(db)
    hoje = date.today().isoformat()
    status           = _calc_estoque_status(db)
    estoque_por_local = _calc_estoque_por_local(db)
    entradas  = db.execute("SELECT * FROM cafe_estoque_entradas ORDER BY data DESC, id DESC LIMIT 60").fetchall()
    contagens = db.execute("SELECT * FROM cafe_estoque_contagens ORDER BY data DESC, id DESC LIMIT 60").fetchall()
    db.close()
    return render_template('cafe_estoque.html',
        status=status, estoque_por_local=estoque_por_local,
        insumos=CAFE_FICHA_INSUMOS,
        locais=CAFE_LOCAIS_TODOS, hoje=hoje,
        entradas=entradas, contagens=contagens,
        alerta_pct=CAFE_ALERTA_PCT, atencao_pct=CAFE_ATENCAO_PCT,
    )


def _pct_para_kg(insumo, qtd_pct):
    """Converte pacotes (PCT) para kg usando o tam_kg da ficha de insumos."""
    info = next((i for i in CAFE_FICHA_INSUMOS if i['nome'] == insumo), None)
    tam_kg = info['tam_kg'] if info else 1.0
    return round(float(qtd_pct) * tam_kg, 4)


@app.route('/cafe/estoque/entrada', methods=['POST'])
def cafe_estoque_entrada():
    db = get_db()
    _init_estoque_tables(db)
    insumo = request.form['insumo']
    qtd_kg = _pct_para_kg(insumo, request.form['qtd_pct'])
    db.execute(
        "INSERT INTO cafe_estoque_entradas (data, origem, local, insumo, qtd_kg, obs) VALUES (?,?,?,?,?,?)",
        (request.form['data'], request.form.get('origem', 'Fornecedor'),
         request.form.get('local', 'Central'),
         insumo, qtd_kg, request.form.get('obs', ''))
    )
    db.commit(); db.close()
    return redirect(url_for('cafe_estoque'))


@app.route('/cafe/estoque/contagem', methods=['POST'])
def cafe_estoque_contagem():
    db = get_db()
    _init_estoque_tables(db)
    insumo = request.form['insumo']
    qtd_kg = _pct_para_kg(insumo, request.form['qtd_pct'])
    db.execute(
        "INSERT INTO cafe_estoque_contagens (data, local, insumo, qtd_kg, obs) VALUES (?,?,?,?,?)",
        (request.form['data'], request.form['local'],
         insumo, qtd_kg, request.form.get('obs', ''))
    )
    db.commit(); db.close()
    return redirect(url_for('cafe_estoque'))


@app.route('/cafe/estoque/del_entrada/<int:eid>', methods=['POST'])
def cafe_del_entrada(eid):
    db = get_db()
    db.execute("DELETE FROM cafe_estoque_entradas WHERE id=?", (eid,))
    db.commit(); db.close()
    return redirect(url_for('cafe_estoque'))


@app.route('/cafe/estoque/del_contagem/<int:cid>', methods=['POST'])
def cafe_del_contagem(cid):
    db = get_db()
    db.execute("DELETE FROM cafe_estoque_contagens WHERE id=?", (cid,))
    db.commit(); db.close()
    return redirect(url_for('cafe_estoque'))

# ─────────────── CAFÉ — EXPORTAR / IMPORTAR ──────────────────────────────────

@app.route('/cafe/exportar')
def cafe_exportar():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    db = get_db()
    _init_estoque_tables(db)

    wb = openpyxl.Workbook()

    # ── Aba 1: Leituras ──
    ws = wb.active
    ws.title = 'Leituras'
    maquinas = db.execute("SELECT * FROM cafe_maquinas ORDER BY id").fetchall()
    maq_map  = {m['id']: m['nome'] for m in maquinas}
    bebidas_cols = [col for col, _ in NESTLE_BEBIDAS]
    bebidas_nomes = [nome for _, nome in NESTLE_BEBIDAS]
    header = ['Máquina', 'PDV', 'Tipo', 'Data'] + bebidas_nomes
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='0F2540')
    rows_l = db.execute("SELECT l.*, m.nome as maq_nome, m.pdv, m.tipo FROM cafe_leituras l JOIN cafe_maquinas m ON l.maquina_id=m.id ORDER BY l.data DESC, l.maquina_id").fetchall()
    for r in rows_l:
        vals = [r[col] for col in bebidas_cols]
        ws.append([r['maq_nome'], r['pdv'], r['tipo'], r['data']] + vals)

    # ── Aba 2: Entradas (compras) ──
    ws2 = wb.create_sheet('Entradas')
    ws2.append(['ID', 'Data', 'Insumo', 'Origem', 'Local', 'Qtd (kg)', 'Obs'])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='375623')
    for r in db.execute("SELECT * FROM cafe_estoque_entradas ORDER BY data DESC").fetchall():
        ws2.append([r['id'], r['data'], r['insumo'], r['origem'], r['local'], r['qtd_kg'], r['obs'] or ''])

    # ── Aba 3: Contagens (inventários) ──
    ws3 = wb.create_sheet('Contagens')
    ws3.append(['ID', 'Data', 'Insumo', 'Local', 'Qtd (kg)'])
    for cell in ws3[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='4A235A')
    for r in db.execute("SELECT * FROM cafe_estoque_contagens ORDER BY data DESC").fetchall():
        ws3.append([r['id'], r['data'], r['insumo'], r['local'], r['qtd_kg']])

    db.close()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from datetime import date as _date
    fname = f'cafe_{_date.today().isoformat()}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/cafe/importar', methods=['POST'])
def cafe_importar():
    import openpyxl
    f = request.files.get('arquivo')
    if not f:
        flash('Nenhum arquivo enviado.', 'danger')
        return redirect(url_for('cafe_lista'))
    aba = request.form.get('aba', 'Tudo')
    wb = openpyxl.load_workbook(f, data_only=True)
    db = get_db()
    _init_estoque_tables(db)
    inseridos = 0

    abas_importar = {'Entradas', 'Contagens', 'Leituras'} if aba == 'Tudo' else {aba}

    if 'Entradas' in abas_importar and 'Entradas' in wb.sheetnames:
        for row in wb['Entradas'].iter_rows(min_row=2, values_only=True):
            _, data, insumo, origem, local, qtd_kg, obs = (list(row) + [None]*7)[:7]
            if not data or not insumo or qtd_kg is None: continue
            data = str(data)[:10]
            if not db.execute("SELECT id FROM cafe_estoque_entradas WHERE data=? AND insumo=? AND qtd_kg=?",
                              (data, insumo, float(qtd_kg))).fetchone():
                db.execute("INSERT INTO cafe_estoque_entradas (data, insumo, origem, local, qtd_kg, obs) VALUES (?,?,?,?,?,?)",
                           (data, insumo, origem or 'Fornecedor', local or 'Central', float(qtd_kg), obs or ''))
                inseridos += 1

    if 'Contagens' in abas_importar and 'Contagens' in wb.sheetnames:
        for row in wb['Contagens'].iter_rows(min_row=2, values_only=True):
            _, data, insumo, local, qtd_kg = (list(row) + [None]*5)[:5]
            if not data or not insumo or qtd_kg is None: continue
            data = str(data)[:10]
            if not db.execute("SELECT id FROM cafe_estoque_contagens WHERE data=? AND insumo=? AND local=?",
                              (data, insumo, local or 'Central')).fetchone():
                db.execute("INSERT INTO cafe_estoque_contagens (data, insumo, local, qtd_kg) VALUES (?,?,?,?)",
                           (data, insumo, local or 'Central', float(qtd_kg)))
                inseridos += 1

    if 'Leituras' in abas_importar and 'Leituras' in wb.sheetnames:
        COLS_L = ['expresso','curto','duplo','americano','capuccino','cafe_leite',
                  'cafe_leite_curto','achoc_kitkat','capp_alpino','achoc_alpino','mocaccino_2f','achoc_2f']
        maq_map = {(r['nome'], r['pdv']): r['id']
                   for r in db.execute("SELECT id, nome, pdv FROM cafe_maquinas").fetchall()}
        for row in wb['Leituras'].iter_rows(min_row=2, values_only=True):
            cells = list(row) + [None]*16
            maq_nome, pdv, tipo, data_str = cells[0], cells[1], cells[2], cells[3]
            vals = cells[4:16]
            if not maq_nome or not data_str: continue
            data_str = str(data_str)[:10]
            mid = maq_map.get((maq_nome, pdv))
            if not mid: continue
            if not db.execute("SELECT id FROM cafe_leituras WHERE maquina_id=? AND data=?", (mid, data_str)).fetchone():
                cols_sql = ', '.join(COLS_L)
                ph = ', '.join(['?']*12)
                db.execute(f"INSERT INTO cafe_leituras (maquina_id, data, {cols_sql}) VALUES (?,?,{ph})",
                           [mid, data_str] + [int(v) if v is not None else None for v in vals])
                inseridos += 1

    db.commit(); db.close()
    flash(f'{inseridos} registro(s) importado(s) com sucesso.', 'success' if inseridos else 'info')
    return redirect(url_for('cafe_estoque'))


# ─────────────────────────────────────────────────────────────────────────────

init_db()

if __name__ == '__main__':
    import webbrowser, threading
    threading.Timer(1.2, lambda: webbrowser.open('http://localhost:5000')).start()
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
